from flask import Flask, render_template, request, jsonify, send_file, session, redirect, url_for
from flask_sqlalchemy import SQLAlchemy
from flask_migrate import Migrate
from dotenv import load_dotenv
from datetime import datetime, date, timedelta
from functools import wraps
import os, io, hashlib, secrets, smtplib, string
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
try:
    from apscheduler.schedulers.background import BackgroundScheduler
    HAS_SCHEDULER = True
except ImportError:
    HAS_SCHEDULER = False

load_dotenv()

app = Flask(__name__)
app.config['SECRET_KEY']               = os.getenv('SECRET_KEY', 'dev-secret-key-mude-isso')
app.config['SQLALCHEMY_DATABASE_URI']  = os.getenv('DATABASE_URL', 'postgresql://postgres:q1w2e3r4@127.0.0.1:5432/patrimonio_db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db      = SQLAlchemy(app)
migrate = Migrate(app, db)

# ══════════════════════════════════════════════════════════════════════════════
# PERFIS E PERMISSÕES
# ══════════════════════════════════════════════════════════════════════════════

PERFIS = {
    'admin':   {'label': 'Administrador', 'pode_ver': True, 'pode_criar': True,  'pode_editar': True,  'pode_excluir': True},
    'tecnico': {'label': 'Técnico',       'pode_ver': True, 'pode_criar': True,  'pode_editar': True,  'pode_excluir': False},
    'comum':   {'label': 'Comum',         'pode_ver': True, 'pode_criar': False, 'pode_editar': False, 'pode_excluir': False},
}

# ══════════════════════════════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def upper(v):
    return v.strip().upper() if v and isinstance(v, str) else v

def _parse_date(s):
    return datetime.strptime(s, '%Y-%m-%d').date()

def to_int(v):
    # Converte para int, retorna None se vazio
    return int(v) if v not in (None, '', 'None') else None

def registrar_log(acao, entidade, detalhe=''):
    try:
        uid = session.get('usuario_id')
        log = LogAtividade(usuario_id=uid, acao=acao, entidade=entidade, detalhe=detalhe[:500])
        db.session.add(log)
    except Exception:
        pass  # log nunca deve quebrar a operacao principal

def registrar_movimentacao(tipo_eq, eq_id, campo, anterior, novo):
    try:
        if str(anterior or '') == str(novo or ''): return
        uid = session.get('usuario_id')
        mov = Movimentacao(tipo_equipamento=tipo_eq, equipamento_id=eq_id,
                           campo_alterado=campo, valor_anterior=str(anterior or ''),
                           valor_novo=str(novo or ''), usuario_id=uid)
        db.session.add(mov)
    except Exception:
        pass


def gerar_senha(n=10):
    chars = string.ascii_letters + string.digits + '!@#$'
    return ''.join(secrets.choice(chars) for _ in range(n))

def hash_senha(senha):
    salt = secrets.token_hex(16)
    h    = hashlib.sha256((salt + senha).encode()).hexdigest()
    return f"{salt}:{h}"

def verificar_senha(senha, senha_hash):
    try:
        salt, h = senha_hash.split(':', 1)
        return hashlib.sha256((salt + senha).encode()).hexdigest() == h
    except Exception:
        return False

def enviar_email(destinatario, assunto, corpo_html):
    """Envia e-mail via SMTP. Retorna (True, '') ou (False, mensagem_erro)."""
    import ssl as _ssl, traceback
    from email.header import Header
    from email.utils import formataddr

    host   = os.getenv('SMTP_HOST', '').strip()
    port   = int(os.getenv('SMTP_PORT', 465))
    user   = os.getenv('SMTP_USER', '').strip()
    passwd = os.getenv('SMTP_PASS', '').strip()
    from_raw = (os.getenv('SMTP_FROM', '') or user).strip()

    if not host:
        msg = "SMTP_HOST nao configurado no .env"
        print(f"[EMAIL] {msg}"); return False, msg
    if not user:
        msg = "SMTP_USER nao configurado no .env"
        print(f"[EMAIL] {msg}"); return False, msg
    if not passwd:
        msg = "SMTP_PASS nao configurado no .env"
        print(f"[EMAIL] {msg}"); return False, msg

    # Codifica o campo From corretamente para suportar caracteres especiais (ex: ô, ã)
    # Extrai nome e endereço se vier no formato "Nome <email>"
    import re
    m = re.match(r'^(.+?)\s*<([^>]+)>$', from_raw)
    if m:
        nome_from  = m.group(1).strip()
        email_from = m.group(2).strip()
        de = formataddr((str(Header(nome_from, 'utf-8')), email_from))
    else:
        de = from_raw

    print(f"[EMAIL] Enviando para {destinatario} via {host}:{port} (user: {user})")
    try:
        msg_obj = MIMEMultipart('alternative')
        msg_obj['Subject'] = Header(assunto, 'utf-8')
        msg_obj['From']    = de
        msg_obj['To']      = destinatario
        msg_obj.attach(MIMEText(corpo_html, 'html', 'utf-8'))

        if port == 465:
            ctx = _ssl.create_default_context()
            with smtplib.SMTP_SSL(host, port, context=ctx, timeout=15) as s:
                s.login(user, passwd)
                s.sendmail(email_from if m else de, [destinatario], msg_obj.as_string())
        else:
            with smtplib.SMTP(host, port, timeout=15) as s:
                s.ehlo(); s.starttls(); s.ehlo()
                s.login(user, passwd)
                s.sendmail(email_from if m else de, [destinatario], msg_obj.as_string())

        print(f"[EMAIL] Enviado com sucesso para {destinatario}")
        return True, ''
    except smtplib.SMTPAuthenticationError as e:
        msg = f"Autenticacao falhou — verifique SMTP_USER e SMTP_PASS"
        print(f"[EMAIL] {msg}: {e}"); return False, msg
    except smtplib.SMTPConnectError as e:
        msg = f"Nao foi possivel conectar a {host}:{port}"
        print(f"[EMAIL] {msg}: {e}"); return False, msg
    except smtplib.SMTPException as e:
        msg = f"Erro SMTP: {e}"
        print(f"[EMAIL] {msg}"); return False, msg
    except OSError as e:
        msg = f"Erro de rede ao conectar a {host}:{port} — {e}"
        print(f"[EMAIL] {msg}"); return False, msg
    except Exception as e:
        msg = f"Erro inesperado: {e}"
        print(f"[EMAIL] {msg}\n{traceback.format_exc()}"); return False, msg

def corpo_email_boas_vindas(nome, email, senha, perfil_label):
    return f"""
    <div style="font-family:Arial,sans-serif;max-width:520px;margin:0 auto;background:#f5f5f5;padding:30px;border-radius:10px;">
      <div style="background:#161B22;border-radius:8px;padding:24px;margin-bottom:20px;">
        <h1 style="color:#2D7DD2;font-size:20px;margin:0;">🖥 PatrimônioTech</h1>
        <p style="color:#8B949E;font-size:12px;margin:4px 0 0;">Sistema de Controle de Patrimônios</p>
      </div>
      <div style="background:#fff;border-radius:8px;padding:24px;">
        <h2 style="color:#1a1a1a;font-size:18px;margin-top:0;">Olá, {nome}!</h2>
        <p style="color:#444;line-height:1.6;">Sua conta foi criada no sistema <strong>PatrimônioTech</strong>. Abaixo estão seus dados de acesso:</p>
        <div style="background:#f0f4f8;border-radius:6px;padding:18px;margin:20px 0;border-left:4px solid #2D7DD2;">
          <table style="width:100%;border-collapse:collapse;">
            <tr><td style="padding:6px 0;color:#666;font-size:13px;">E-mail:</td><td style="padding:6px 0;font-weight:bold;font-size:13px;">{email}</td></tr>
            <tr><td style="padding:6px 0;color:#666;font-size:13px;">Senha:</td><td style="padding:6px 0;font-weight:bold;font-size:13px;font-family:monospace;background:#e8edf2;padding:4px 8px;border-radius:4px;">{senha}</td></tr>
            <tr><td style="padding:6px 0;color:#666;font-size:13px;">Perfil:</td><td style="padding:6px 0;font-size:13px;">{perfil_label}</td></tr>
          </table>
        </div>
        <p style="color:#e74c3c;font-size:12px;">⚠ Por segurança, recomendamos alterar a senha no primeiro acesso.</p>
        <p style="color:#666;font-size:12px;margin-top:20px;">Em caso de dúvidas, entre em contato com o administrador do sistema.</p>
      </div>
    </div>"""

# ══════════════════════════════════════════════════════════════════════════════
# DECORATORS
# ══════════════════════════════════════════════════════════════════════════════

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'usuario_id' not in session:
            if request.is_json:
                return jsonify({'erro': 'Não autenticado', 'redirect': '/login'}), 401
            return redirect(url_for('login_page'))
        return f(*args, **kwargs)
    return decorated

def perfil_required(*perfis_permitidos):
    def decorator(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            if 'usuario_id' not in session:
                return jsonify({'erro': 'Não autenticado'}), 401
            if session.get('perfil') not in perfis_permitidos:
                return jsonify({'erro': 'Sem permissão para esta ação', 'sem_permissao': True}), 403
            return f(*args, **kwargs)
        return decorated
    return decorator

# ══════════════════════════════════════════════════════════════════════════════
# MODELS
# ══════════════════════════════════════════════════════════════════════════════

class Usuario(db.Model):
    __tablename__ = 'usuarios'
    id            = db.Column(db.Integer, primary_key=True)
    nome          = db.Column(db.String(120), nullable=False)
    email         = db.Column(db.String(150), nullable=False, unique=True)
    senha_hash    = db.Column(db.String(200), nullable=False)
    perfil        = db.Column(db.String(20),  nullable=False, default='comum')
    ativo           = db.Column(db.Boolean, default=True)
    senha_alterada  = db.Column(db.Boolean, default=False)  # False = primeiro acesso
    criado_em       = db.Column(db.DateTime, default=datetime.utcnow)
    ultimo_acesso   = db.Column(db.DateTime)

    def to_dict(self):
        return {
            'id': self.id, 'nome': self.nome, 'email': self.email,
            'perfil': self.perfil,
            'perfil_label': PERFIS.get(self.perfil, {}).get('label', self.perfil),
            'ativo': self.ativo,
            'senha_alterada': self.senha_alterada,
            'criado_em':     self.criado_em.strftime('%d/%m/%Y')       if self.criado_em     else '',
            'ultimo_acesso': self.ultimo_acesso.strftime('%d/%m/%Y %H:%M') if self.ultimo_acesso else 'Nunca',
        }


class Empresa(db.Model):
    __tablename__ = 'empresas'
    id        = db.Column(db.Integer, primary_key=True)
    nome      = db.Column(db.String(150), nullable=False, unique=True)
    cnpj      = db.Column(db.String(20))
    contato   = db.Column(db.String(100))
    telefone  = db.Column(db.String(20))
    criado_em = db.Column(db.DateTime, default=datetime.utcnow)
    monitores       = db.relationship('Monitor',       backref='empresa', lazy=True)
    desktops        = db.relationship('Desktop',       backref='empresa', lazy=True)
    estabilizadores = db.relationship('Estabilizador', backref='empresa', lazy=True)
    notebooks       = db.relationship('Notebook',      backref='empresa', lazy=True)

    def to_dict(self):
        return {'id': self.id, 'nome': self.nome, 'cnpj': self.cnpj,
                'contato': self.contato, 'telefone': self.telefone}


class Localizacao(db.Model):
    __tablename__ = 'localizacoes'
    id        = db.Column(db.Integer, primary_key=True)
    nome      = db.Column(db.String(200), nullable=False, unique=True)
    descricao = db.Column(db.String(300))
    criado_em = db.Column(db.DateTime, default=datetime.utcnow)

    def to_dict(self):
        return {'id': self.id, 'nome': self.nome, 'descricao': self.descricao}


class ConfigAlerta(db.Model):
    __tablename__ = 'config_alerta'
    id           = db.Column(db.Integer, primary_key=True)
    emails       = db.Column(db.Text, default='')
    dias_aviso   = db.Column(db.Integer, default=10)
    ativo        = db.Column(db.Boolean, default=True)
    ultimo_envio = db.Column(db.DateTime)
    def get_emails(self):
        return [e.strip() for e in (self.emails or '').split(',') if e.strip()]

class Chamado(db.Model):
    __tablename__ = 'chamados'
    id                = db.Column(db.Integer, primary_key=True)
    numero_chamado    = db.Column(db.String(100), nullable=False)
    tipo_equipamento  = db.Column(db.String(30), nullable=False)  # Monitor/Desktop/Estabilizador/Notebook
    equipamento_id    = db.Column(db.Integer, nullable=False)
    data_abertura     = db.Column(db.Date, nullable=False)
    data_solucao      = db.Column(db.Date)
    descricao         = db.Column(db.Text)
    solucao           = db.Column(db.Text)
    status            = db.Column(db.String(20), default='aberto')  # aberto / fechado
    prioridade        = db.Column(db.String(20), default='media')   # baixa/media/alta/critica
    status_detalhe    = db.Column(db.String(30), default='em_analise')  # em_analise/aguardando_peca/em_reparo/resolvido
    criado_em         = db.Column(db.DateTime, default=datetime.utcnow)

    def to_dict(self):
        return {
            'id':               self.id,
            'numero_chamado':   self.numero_chamado,
            'tipo_equipamento': self.tipo_equipamento,
            'equipamento_id':   self.equipamento_id,
            'data_abertura':    self.data_abertura.strftime('%d/%m/%Y') if self.data_abertura else '',
            'data_solucao':     self.data_solucao.strftime('%d/%m/%Y') if self.data_solucao else '',
            'descricao':        self.descricao,
            'solucao':          self.solucao,
            'status':           self.status,
            'prioridade':       self.prioridade or 'media',
            'status_detalhe':   self.status_detalhe or 'em_analise',
            'criado_em':        self.criado_em.strftime('%d/%m/%Y') if self.criado_em else '',
        }


class Movimentacao(db.Model):
    __tablename__ = 'movimentacoes'
    id               = db.Column(db.Integer, primary_key=True)
    tipo_equipamento = db.Column(db.String(30), nullable=False)
    equipamento_id   = db.Column(db.Integer, nullable=False)
    campo_alterado   = db.Column(db.String(50))   # localizacao / empresa / tipo / etc
    valor_anterior   = db.Column(db.Text)
    valor_novo       = db.Column(db.Text)
    usuario_id       = db.Column(db.Integer, db.ForeignKey('usuarios.id'))
    usuario          = db.relationship('Usuario', foreign_keys=[usuario_id])
    criado_em        = db.Column(db.DateTime, default=datetime.utcnow)
    def to_dict(self):
        return {
            'id': self.id,
            'tipo_equipamento': self.tipo_equipamento,
            'equipamento_id':   self.equipamento_id,
            'campo_alterado':   self.campo_alterado,
            'valor_anterior':   self.valor_anterior,
            'valor_novo':       self.valor_novo,
            'usuario':          self.usuario.nome if self.usuario else '—',
            'criado_em':        self.criado_em.strftime('%d/%m/%Y %H:%M') if self.criado_em else '',
        }

class LogAtividade(db.Model):
    __tablename__ = 'log_atividades'
    id         = db.Column(db.Integer, primary_key=True)
    usuario_id = db.Column(db.Integer, db.ForeignKey('usuarios.id'))
    usuario    = db.relationship('Usuario', foreign_keys=[usuario_id])
    acao       = db.Column(db.String(50))   # criar / editar / excluir / login / etc
    entidade   = db.Column(db.String(50))   # Monitor / Desktop / Chamado / etc
    detalhe    = db.Column(db.Text)
    criado_em  = db.Column(db.DateTime, default=datetime.utcnow)
    def to_dict(self):
        return {
            'id':        self.id,
            'usuario':   self.usuario.nome if self.usuario else '—',
            'acao':      self.acao,
            'entidade':  self.entidade,
            'detalhe':   self.detalhe,
            'criado_em': self.criado_em.strftime('%d/%m/%Y %H:%M') if self.criado_em else '',
        }

class ComentarioChamado(db.Model):
    __tablename__ = 'comentarios_chamado'
    id         = db.Column(db.Integer, primary_key=True)
    chamado_id = db.Column(db.Integer, db.ForeignKey('chamados.id'), nullable=False)
    usuario_id = db.Column(db.Integer, db.ForeignKey('usuarios.id'))
    usuario    = db.relationship('Usuario', foreign_keys=[usuario_id])
    texto      = db.Column(db.Text, nullable=False)
    criado_em  = db.Column(db.DateTime, default=datetime.utcnow)
    def to_dict(self):
        return {
            'id':        self.id,
            'chamado_id': self.chamado_id,
            'usuario':   self.usuario.nome if self.usuario else '—',
            'texto':     self.texto,
            'criado_em': self.criado_em.strftime('%d/%m/%Y %H:%M') if self.criado_em else '',
        }

class Monitor(db.Model):
    __tablename__ = 'monitores'
    id                  = db.Column(db.Integer, primary_key=True)
    numero_patrimonio   = db.Column(db.String(50), nullable=False, unique=True)
    pa                  = db.Column(db.String(50))
    marca               = db.Column(db.String(80))
    modelo              = db.Column(db.String(80))
    tamanho_polegadas   = db.Column(db.Float)
    localizacao         = db.Column(db.String(200), nullable=False)
    empresa_id          = db.Column(db.Integer, db.ForeignKey('empresas.id'), nullable=False)
    data_aluguel        = db.Column(db.Date, nullable=False)
    data_fim_fidelidade = db.Column(db.Date, nullable=False)
    observacoes         = db.Column(db.Text)
    ativo               = db.Column(db.Boolean, default=True)
    criado_em           = db.Column(db.DateTime, default=datetime.utcnow)
    atualizado_em       = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    def dias_para_vencer(self): return (self.data_fim_fidelidade - date.today()).days if self.data_fim_fidelidade else None
    def to_dict(self):
        return {'id': self.id, 'tipo': 'Monitor', 'numero_patrimonio': self.numero_patrimonio, 'pa': self.pa,
                'marca': self.marca, 'modelo': self.modelo, 'tamanho_polegadas': self.tamanho_polegadas,
                'localizacao': self.localizacao, 'empresa': self.empresa.nome if self.empresa else '',
                'empresa_id': self.empresa_id,
                'data_aluguel': self.data_aluguel.strftime('%d/%m/%Y') if self.data_aluguel else '',
                'data_fim_fidelidade': self.data_fim_fidelidade.strftime('%d/%m/%Y') if self.data_fim_fidelidade else '',
                'observacoes': self.observacoes, 'ativo': self.ativo,
            'criado_em': self.criado_em.strftime('%d/%m/%Y') if self.criado_em else ''}


class Desktop(db.Model):
    __tablename__ = 'desktops'
    id                  = db.Column(db.Integer, primary_key=True)
    numero_patrimonio   = db.Column(db.String(50), nullable=False, unique=True)
    pa                  = db.Column(db.String(50))
    marca               = db.Column(db.String(80))
    modelo              = db.Column(db.String(80))
    processador         = db.Column(db.String(100))
    ram_gb              = db.Column(db.Integer)
    armazenamento_gb    = db.Column(db.Integer)
    localizacao         = db.Column(db.String(200), nullable=False)
    empresa_id          = db.Column(db.Integer, db.ForeignKey('empresas.id'), nullable=False)
    data_aluguel        = db.Column(db.Date, nullable=False)
    data_fim_fidelidade = db.Column(db.Date, nullable=False)
    observacoes         = db.Column(db.Text)
    ativo               = db.Column(db.Boolean, default=True)
    criado_em           = db.Column(db.DateTime, default=datetime.utcnow)
    atualizado_em       = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    def dias_para_vencer(self): return (self.data_fim_fidelidade - date.today()).days if self.data_fim_fidelidade else None
    def to_dict(self):
        return {'id': self.id, 'tipo': 'Desktop', 'numero_patrimonio': self.numero_patrimonio, 'pa': self.pa,
                'marca': self.marca, 'modelo': self.modelo, 'processador': self.processador,
                'ram_gb': self.ram_gb, 'armazenamento_gb': self.armazenamento_gb,
                'localizacao': self.localizacao, 'empresa': self.empresa.nome if self.empresa else '',
                'empresa_id': self.empresa_id,
                'data_aluguel': self.data_aluguel.strftime('%d/%m/%Y') if self.data_aluguel else '',
                'data_fim_fidelidade': self.data_fim_fidelidade.strftime('%d/%m/%Y') if self.data_fim_fidelidade else '',
                'observacoes': self.observacoes, 'ativo': self.ativo,
                'criado_em': self.criado_em.strftime('%d/%m/%Y') if self.criado_em else ''}


class Estabilizador(db.Model):
    __tablename__ = 'estabilizadores'
    id                  = db.Column(db.Integer, primary_key=True)
    numero_patrimonio   = db.Column(db.String(50), nullable=False, unique=True)
    pa                  = db.Column(db.String(50))
    marca               = db.Column(db.String(80))
    modelo              = db.Column(db.String(80))
    potencia_va         = db.Column(db.Integer)
    localizacao         = db.Column(db.String(200), nullable=False)
    empresa_id          = db.Column(db.Integer, db.ForeignKey('empresas.id'), nullable=False)
    data_aluguel        = db.Column(db.Date, nullable=False)
    data_fim_fidelidade = db.Column(db.Date, nullable=False)
    observacoes         = db.Column(db.Text)
    ativo               = db.Column(db.Boolean, default=True)
    criado_em           = db.Column(db.DateTime, default=datetime.utcnow)
    atualizado_em       = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    def dias_para_vencer(self): return (self.data_fim_fidelidade - date.today()).days if self.data_fim_fidelidade else None
    def to_dict(self):
        return {'id': self.id, 'tipo': 'Estabilizador', 'numero_patrimonio': self.numero_patrimonio, 'pa': self.pa,
                'marca': self.marca, 'modelo': self.modelo, 'potencia_va': self.potencia_va,
                'localizacao': self.localizacao, 'empresa': self.empresa.nome if self.empresa else '',
                'empresa_id': self.empresa_id,
                'data_aluguel': self.data_aluguel.strftime('%d/%m/%Y') if self.data_aluguel else '',
                'data_fim_fidelidade': self.data_fim_fidelidade.strftime('%d/%m/%Y') if self.data_fim_fidelidade else '',
                'observacoes': self.observacoes, 'ativo': self.ativo,
                'criado_em': self.criado_em.strftime('%d/%m/%Y') if self.criado_em else ''}


class Notebook(db.Model):
    __tablename__ = 'notebooks'
    id                  = db.Column(db.Integer, primary_key=True)
    numero_patrimonio   = db.Column(db.String(50), nullable=False, unique=True)
    pa                  = db.Column(db.String(50))
    marca               = db.Column(db.String(80))
    modelo              = db.Column(db.String(80))
    processador         = db.Column(db.String(100))
    ram_gb              = db.Column(db.Integer)
    armazenamento_gb    = db.Column(db.Integer)
    tamanho_tela        = db.Column(db.Float)
    localizacao         = db.Column(db.String(200), nullable=False)
    empresa_id          = db.Column(db.Integer, db.ForeignKey('empresas.id'), nullable=False)
    data_aluguel        = db.Column(db.Date, nullable=False)
    data_fim_fidelidade = db.Column(db.Date, nullable=False)
    observacoes         = db.Column(db.Text)
    ativo               = db.Column(db.Boolean, default=True)
    criado_em           = db.Column(db.DateTime, default=datetime.utcnow)
    atualizado_em       = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    def dias_para_vencer(self): return (self.data_fim_fidelidade - date.today()).days if self.data_fim_fidelidade else None
    def to_dict(self):
        return {'id': self.id, 'tipo': 'Notebook', 'numero_patrimonio': self.numero_patrimonio, 'pa': self.pa,
                'marca': self.marca, 'modelo': self.modelo, 'processador': self.processador,
                'ram_gb': self.ram_gb, 'armazenamento_gb': self.armazenamento_gb,
                'tamanho_tela': self.tamanho_tela,
                'localizacao': self.localizacao, 'empresa': self.empresa.nome if self.empresa else '',
                'empresa_id': self.empresa_id,
                'data_aluguel': self.data_aluguel.strftime('%d/%m/%Y') if self.data_aluguel else '',
                'data_fim_fidelidade': self.data_fim_fidelidade.strftime('%d/%m/%Y') if self.data_fim_fidelidade else '',
                'observacoes': self.observacoes, 'ativo': self.ativo,
                'criado_em': self.criado_em.strftime('%d/%m/%Y') if self.criado_em else ''}

# ══════════════════════════════════════════════════════════════════════════════
# ROTAS — AUTH
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/login')
def login_page():
    if 'usuario_id' in session:
        return redirect(url_for('index'))
    return render_template('login.html')

@app.route('/api/login', methods=['POST'])
def fazer_login():
    data  = request.json
    email = (data.get('email') or '').strip().lower()
    senha = data.get('senha') or ''
    u = Usuario.query.filter_by(email=email, ativo=True).first()
    if not u or not verificar_senha(senha, u.senha_hash):
        return jsonify({'sucesso': False, 'erro': 'E-mail ou senha incorretos'}), 401
    u.ultimo_acesso = datetime.utcnow()
    db.session.commit()
    session.permanent = True
    session['usuario_id'] = u.id
    session['nome']        = u.nome
    session['perfil']      = u.perfil
    perms = PERFIS.get(u.perfil, {})
    session['pode_criar']    = perms.get('pode_criar',   False)
    session['pode_editar']   = perms.get('pode_editar',  False)
    session['pode_excluir']  = perms.get('pode_excluir', False)
    session['senha_alterada'] = u.senha_alterada
    session['email'] = u.email
    return jsonify({'sucesso': True, 'perfil': u.perfil, 'nome': u.nome, 'senha_alterada': u.senha_alterada})

@app.route('/api/logout', methods=['POST'])
def logout():
    session.clear()
    return jsonify({'sucesso': True})

@app.route('/api/me')
@login_required
def me():
    return jsonify({
        'nome':           session.get('nome'),
        'email':          session.get('email',''),
        'perfil':         session.get('perfil'),
        'perfil_label':   PERFIS.get(session.get('perfil',''), {}).get('label',''),
        'pode_criar':     session.get('pode_criar',   False),
        'pode_editar':    session.get('pode_editar',  False),
        'pode_excluir':   session.get('pode_excluir', False),
        'senha_alterada': session.get('senha_alterada', True),
    })

@app.route('/api/me', methods=['PUT'])
@login_required
def atualizar_perfil():
    data = request.json
    uid  = session.get('usuario_id')
    u    = db.get_or_404(Usuario, uid)
    try:
        novo_nome  = (data.get('nome') or '').strip()
        novo_email = (data.get('email') or '').strip().lower()
        if not novo_nome or not novo_email:
            return jsonify({'sucesso': False, 'erro': 'Nome e e-mail são obrigatórios'}), 400
        # Verifica se e-mail já existe em outro usuário
        outro = Usuario.query.filter_by(email=novo_email).first()
        if outro and outro.id != uid:
            return jsonify({'sucesso': False, 'erro': 'E-mail já em uso por outro usuário'}), 400
        u.nome  = novo_nome
        u.email = novo_email
        session['nome']  = novo_nome
        session['email'] = novo_email
        registrar_log('editar', 'Perfil', novo_email)
        db.session.commit()
        return jsonify({'sucesso': True})
    except Exception as ex:
        db.session.rollback()
        return jsonify({'sucesso': False, 'erro': str(ex)[:150]}), 400

# ══════════════════════════════════════════════════════════════════════════════
# ROTAS — USUÁRIOS (admin only)
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/api/usuarios', methods=['GET'])
@perfil_required('admin')
def listar_usuarios():
    return jsonify([u.to_dict() for u in Usuario.query.order_by(Usuario.nome).all()])

@app.route('/api/usuarios', methods=['POST'])
@perfil_required('admin')
def criar_usuario():
    data  = request.json
    nome  = (data.get('nome') or '').strip()
    email = (data.get('email') or '').strip().lower()
    perfil = data.get('perfil', 'comum')
    if not nome or not email:
        return jsonify({'sucesso': False, 'erro': 'Nome e e-mail são obrigatórios'}), 400
    if perfil not in PERFIS:
        return jsonify({'sucesso': False, 'erro': 'Perfil inválido'}), 400
    if Usuario.query.filter_by(email=email).first():
        return jsonify({'sucesso': False, 'erro': 'E-mail já cadastrado'}), 400
    senha_gerada = gerar_senha()
    u = Usuario(nome=nome, email=email, perfil=perfil,
                senha_hash=hash_senha(senha_gerada))
    db.session.add(u); db.session.commit()
    perfil_label = PERFIS[perfil]['label']
    email_ok, email_erro = enviar_email(
        email,
        '🖥 PatrimônioTech — Seus dados de acesso',
        corpo_email_boas_vindas(nome, email, senha_gerada, perfil_label)
    )
    return jsonify({
        'sucesso': True, 'id': u.id,
        'email_enviado': email_ok,
        'email_erro': email_erro,
        'senha': senha_gerada
    }), 201

@app.route('/api/usuarios/<int:id>', methods=['PUT'])
@perfil_required('admin')
def atualizar_usuario(id):
    u = db.get_or_404(Usuario, id); data = request.json
    if 'nome' in data:   u.nome   = data['nome'].strip()
    if 'email' in data:  u.email  = data['email'].strip().lower()
    if 'perfil' in data and data['perfil'] in PERFIS: u.perfil = data['perfil']
    if 'ativo' in data:  u.ativo  = bool(data['ativo'])
    if data.get('resetar_senha'):
        nova_senha = gerar_senha()
        u.senha_hash = hash_senha(nova_senha)
        u.senha_alterada = False  # força troca no próximo login
        db.session.commit()
        email_ok, email_erro = enviar_email(u.email, '🖥 PatrimônioTech — Senha redefinida',
            corpo_email_boas_vindas(u.nome, u.email, nova_senha, PERFIS[u.perfil]['label']))
        return jsonify({'sucesso': True, 'senha': nova_senha, 'email_enviado': email_ok, 'email_erro': email_erro})
    if data.get('definir_senha'):
        nova = data['definir_senha']
        if len(nova) < 6:
            return jsonify({'sucesso': False, 'erro': 'Senha deve ter ao menos 6 caracteres'}), 400
        u.senha_hash = hash_senha(nova)
        u.senha_alterada = False  # força troca no próximo login
        db.session.commit()
        return jsonify({'sucesso': True})
    try:
        db.session.commit(); return jsonify({'sucesso': True})
    except Exception as ex:
        db.session.rollback(); return jsonify({'sucesso': False, 'erro': str(ex)}), 400

@app.route('/api/usuarios/<int:id>', methods=['DELETE'])
@perfil_required('admin')
def deletar_usuario(id):
    u = db.get_or_404(Usuario, id)
    if u.id == session.get('usuario_id'):
        return jsonify({'sucesso': False, 'erro': 'Não é possível excluir o próprio usuário'}), 400
    u.ativo = False; db.session.commit()
    return jsonify({'sucesso': True})

@app.route('/api/alterar_senha', methods=['POST'])
@login_required
def alterar_senha():
    data    = request.json
    atual   = data.get('senha_atual', '')
    nova    = data.get('senha_nova', '')
    u = db.get_or_404(Usuario, session['usuario_id'])
    if not verificar_senha(atual, u.senha_hash):
        return jsonify({'sucesso': False, 'erro': 'Senha atual incorreta'}), 400
    if len(nova) < 6:
        return jsonify({'sucesso': False, 'erro': 'A nova senha deve ter ao menos 6 caracteres'}), 400
    u.senha_hash = hash_senha(nova)
    u.senha_alterada = True
    session['senha_alterada'] = True
    db.session.commit()
    return jsonify({'sucesso': True})

# ══════════════════════════════════════════════════════════════════════════════
# ROTA PRINCIPAL
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/trocar-senha')
@login_required
def trocar_senha_page():
    return render_template('trocar_senha.html',
        usuario_nome=session.get('nome',''),
    )

@app.route('/')
@login_required
def index():
    # Força troca de senha no primeiro acesso
    if not session.get('senha_alterada', True):
        return redirect(url_for('trocar_senha_page'))
    hoje   = date.today()
    limite = date.fromordinal(hoje.toordinal() + 30)
    alertas = sum([
        Monitor.query.filter(Monitor.ativo==True,           Monitor.data_fim_fidelidade<=limite).count(),
        Desktop.query.filter(Desktop.ativo==True,           Desktop.data_fim_fidelidade<=limite).count(),
        Estabilizador.query.filter(Estabilizador.ativo==True, Estabilizador.data_fim_fidelidade<=limite).count(),
        Notebook.query.filter(Notebook.ativo==True,         Notebook.data_fim_fidelidade<=limite).count(),
    ])
    return render_template('index.html',
        total_monitores      = Monitor.query.filter_by(ativo=True).count(),
        total_desktops       = Desktop.query.filter_by(ativo=True).count(),
        total_estabilizadores= Estabilizador.query.filter_by(ativo=True).count(),
        total_notebooks      = Notebook.query.filter_by(ativo=True).count(),
        total_empresas       = Empresa.query.count(),
        alertas=alertas,
        usuario_nome   = session.get('nome',''),
        usuario_perfil = session.get('perfil',''),
        perfil_label   = PERFIS.get(session.get('perfil',''),{}).get('label',''),
        pode_criar     = session.get('pode_criar',  False),
        pode_editar    = session.get('pode_editar', False),
        pode_excluir   = session.get('pode_excluir',False),
    )

# ══════════════════════════════════════════════════════════════════════════════
# ROTAS — EQUIPAMENTOS
# ══════════════════════════════════════════════════════════════════════════════

def _filtrar(model, q):
    busca      = request.args.get('busca', '').strip()
    empresa_id = request.args.get('empresa_id', '').strip()
    if busca:
        q = q.filter(db.or_(
            model.numero_patrimonio.ilike(f'%{busca}%'),
            model.localizacao.ilike(f'%{busca}%'),
            model.modelo.ilike(f'%{busca}%'),
        ))
    if empresa_id:
        q = q.filter_by(empresa_id=empresa_id)
    return q

@app.route('/api/stats')
@login_required
def get_stats():
    hoje   = date.today()
    limite = date.fromordinal(hoje.toordinal() + 30)
    alertas = sum([
        Monitor.query.filter(Monitor.ativo==True, Monitor.data_fim_fidelidade<=limite).count(),
        Desktop.query.filter(Desktop.ativo==True, Desktop.data_fim_fidelidade<=limite).count(),
        Estabilizador.query.filter(Estabilizador.ativo==True, Estabilizador.data_fim_fidelidade<=limite).count(),
        Notebook.query.filter(Notebook.ativo==True, Notebook.data_fim_fidelidade<=limite).count(),
    ])
    return jsonify({
        'total_monitores':       Monitor.query.filter_by(ativo=True).count(),
        'total_desktops':        Desktop.query.filter_by(ativo=True).count(),
        'total_estabilizadores': Estabilizador.query.filter_by(ativo=True).count(),
        'total_notebooks':       Notebook.query.filter_by(ativo=True).count(),
        'total_empresas':        Empresa.query.count(),
        'alertas':               alertas,
    })

@app.route('/api/dashboard')
@login_required
def dashboard_equipamentos():
    from itertools import chain
    page  = int(request.args.get('page', 1))
    limit = int(request.args.get('limit', 20))
    monitores       = Monitor.query.filter_by(ativo=True).all()
    desktops        = Desktop.query.filter_by(ativo=True).all()
    estabilizadores = Estabilizador.query.filter_by(ativo=True).all()
    notebooks       = Notebook.query.filter_by(ativo=True).all()
    todos = sorted(
        chain(monitores, desktops, estabilizadores, notebooks),
        key=lambda x: x.data_fim_fidelidade or date.max,
        reverse=False
    )
    total = len(todos)
    pages = max(1, -(-total // limit))
    items = todos[(page-1)*limit : page*limit]
    return jsonify({'items': [e.to_dict() for e in items],
                    'total': total, 'page': page, 'pages': pages})

@app.route('/api/equipamentos')
@login_required
def listar_equipamentos():
    tipo   = request.args.get('tipo', 'todos')
    alerta = request.args.get('alerta', '')
    page   = int(request.args.get('page', 1))
    limit  = int(request.args.get('limit', 20))  # 20 por página
    res = []
    if tipo in ('todos','monitor'):      res += [m.to_dict() for m in _filtrar(Monitor,       Monitor.query.filter_by(ativo=True)).all()]
    if tipo in ('todos','desktop'):      res += [d.to_dict() for d in _filtrar(Desktop,       Desktop.query.filter_by(ativo=True)).all()]
    if tipo in ('todos','estabilizador'):res += [e.to_dict() for e in _filtrar(Estabilizador, Estabilizador.query.filter_by(ativo=True)).all()]
    if tipo in ('todos','notebook'):     res += [n.to_dict() for n in _filtrar(Notebook,      Notebook.query.filter_by(ativo=True)).all()]
    if alerta == '1':
        limite = date.today() + timedelta(days=30)
        res = [e for e in res if e.get('data_fim_fidelidade') and
               datetime.strptime(e['data_fim_fidelidade'], '%d/%m/%Y').date() <= limite]
        res.sort(key=lambda e: datetime.strptime(e['data_fim_fidelidade'], '%d/%m/%Y').date())
    total = len(res)
    # Paginação
    if limit > 0:
        res = res[(page-1)*limit : page*limit]
    return jsonify({'items': res, 'total': total, 'page': page,
                    'pages': max(1, -(-total // limit)) if limit > 0 else 1})

@app.route('/api/patrimonio/<codigo>')
@login_required
def buscar_patrimonio(codigo):
    for model in [Monitor, Desktop, Estabilizador, Notebook]:
        obj = model.query.filter_by(numero_patrimonio=codigo, ativo=True).first()
        if obj: return jsonify({'encontrado': True, 'dados': obj.to_dict()})
    return jsonify({'encontrado': False})

# ── Monitores ──────────────────────────────────────────────────────────────────
@app.route('/api/monitores', methods=['POST'])
@perfil_required('admin','tecnico')
def criar_monitor():
    data = request.json
    try:
        pat = upper(data['numero_patrimonio'])
        m = Monitor.query.filter_by(numero_patrimonio=pat).first()
        if m:
            m.ativo = True
        else:
            m = Monitor(); db.session.add(m)
        m.numero_patrimonio   = pat
        m.pa                  = upper(data.get('pa'))
        m.marca               = upper(data.get('marca'))
        m.modelo              = upper(data.get('modelo'))
        m.tamanho_polegadas   = to_int(data.get('tamanho_polegadas'))
        m.localizacao         = upper(data['localizacao'])
        m.empresa_id          = to_int(data['empresa_id'])
        m.data_aluguel        = _parse_date(data['data_aluguel'])
        m.data_fim_fidelidade = _parse_date(data['data_fim_fidelidade'])
        m.observacoes         = upper(data.get('observacoes'))
        db.session.commit()
        return jsonify({'sucesso': True, 'id': m.id}), 201
    except Exception as e: db.session.rollback(); return jsonify({'sucesso': False, 'erro': str(e)[:150]}), 400

@app.route('/api/monitores/<int:id>', methods=['PUT'])
@perfil_required('admin','tecnico')
def atualizar_monitor(id):
    m = db.get_or_404(Monitor, id); data = request.json
    try:
        registrar_movimentacao('Monitor', id, 'localizacao', m.localizacao, upper(data.get('localizacao', m.localizacao)))
        emp_ant = m.empresa.nome if m.empresa else ''
        nova_emp = Empresa.query.get(to_int(data.get('empresa_id', m.empresa_id)))
        registrar_movimentacao('Monitor', id, 'empresa', emp_ant, nova_emp.nome if nova_emp else '')
        m.numero_patrimonio=upper(data.get('numero_patrimonio',m.numero_patrimonio)); m.marca=upper(data.get('marca',m.marca)); m.pa=upper(data.get('pa',m.pa))
        m.modelo=upper(data.get('modelo',m.modelo)); m.tamanho_polegadas=data.get('tamanho_polegadas',m.tamanho_polegadas)
        m.localizacao=upper(data.get('localizacao',m.localizacao)); m.empresa_id=data.get('empresa_id',m.empresa_id)
        m.observacoes=upper(data.get('observacoes',m.observacoes))
        if data.get('data_aluguel'):        m.data_aluguel=_parse_date(data['data_aluguel'])
        if data.get('data_fim_fidelidade'): m.data_fim_fidelidade=_parse_date(data['data_fim_fidelidade'])
        registrar_log('editar', 'Monitor', m.numero_patrimonio)
        db.session.commit(); return jsonify({'sucesso': True})
    except Exception as e: db.session.rollback(); return jsonify({'sucesso': False, 'erro': str(e)[:150]}), 400

@app.route('/api/monitores/<int:id>', methods=['DELETE'])
@perfil_required('admin')
def deletar_monitor(id):
    m = db.get_or_404(Monitor, id); m.ativo=False; db.session.commit(); return jsonify({'sucesso': True})

# ── Desktops ───────────────────────────────────────────────────────────────────
@app.route('/api/desktops', methods=['POST'])
@perfil_required('admin','tecnico')
def criar_desktop():
    data = request.json
    try:
        pat = upper(data['numero_patrimonio'])
        d = Desktop.query.filter_by(numero_patrimonio=pat).first()
        if d:
            d.ativo = True
        else:
            d = Desktop(); db.session.add(d)
        d.numero_patrimonio   = pat
        d.pa                  = upper(data.get('pa'))
        d.marca               = upper(data.get('marca'))
        d.modelo              = upper(data.get('modelo'))
        d.processador         = upper(data.get('processador'))
        d.ram_gb              = to_int(data.get('ram_gb'))
        d.armazenamento_gb    = to_int(data.get('armazenamento_gb'))
        d.localizacao         = upper(data['localizacao'])
        d.empresa_id          = to_int(data['empresa_id'])
        d.data_aluguel        = _parse_date(data['data_aluguel'])
        d.data_fim_fidelidade = _parse_date(data['data_fim_fidelidade'])
        d.observacoes         = upper(data.get('observacoes'))
        db.session.commit()
        return jsonify({'sucesso': True, 'id': d.id}), 201
    except Exception as e: db.session.rollback(); return jsonify({'sucesso': False, 'erro': str(e)[:150]}), 400

@app.route('/api/desktops/<int:id>', methods=['PUT'])
@perfil_required('admin','tecnico')
def atualizar_desktop(id):
    d = db.get_or_404(Desktop, id); data = request.json
    try:
        registrar_movimentacao('Desktop', id, 'localizacao', d.localizacao, upper(data.get('localizacao', d.localizacao)))
        emp_ant = d.empresa.nome if d.empresa else ''
        nova_emp = Empresa.query.get(to_int(data.get('empresa_id', d.empresa_id)))
        registrar_movimentacao('Desktop', id, 'empresa', emp_ant, nova_emp.nome if nova_emp else '')
        d.numero_patrimonio=upper(data.get('numero_patrimonio',d.numero_patrimonio)); d.marca=upper(data.get('marca',d.marca)); d.pa=upper(data.get('pa',d.pa))
        d.modelo=upper(data.get('modelo',d.modelo)); d.processador=upper(data.get('processador',d.processador))
        d.ram_gb=data.get('ram_gb',d.ram_gb); d.armazenamento_gb=data.get('armazenamento_gb',d.armazenamento_gb)
        d.localizacao=upper(data.get('localizacao',d.localizacao)); d.empresa_id=data.get('empresa_id',d.empresa_id)
        d.observacoes=upper(data.get('observacoes',d.observacoes))
        if data.get('data_aluguel'):        d.data_aluguel=_parse_date(data['data_aluguel'])
        if data.get('data_fim_fidelidade'): d.data_fim_fidelidade=_parse_date(data['data_fim_fidelidade'])
        registrar_log('editar', 'Desktop', d.numero_patrimonio)
        db.session.commit(); return jsonify({'sucesso': True})
    except Exception as e: db.session.rollback(); return jsonify({'sucesso': False, 'erro': str(e)[:150]}), 400

@app.route('/api/desktops/<int:id>', methods=['DELETE'])
@perfil_required('admin')
def deletar_desktop(id):
    d = db.get_or_404(Desktop, id); d.ativo=False; db.session.commit(); return jsonify({'sucesso': True})

# ── Estabilizadores ────────────────────────────────────────────────────────────
@app.route('/api/estabilizadores', methods=['POST'])
@perfil_required('admin','tecnico')
def criar_estabilizador():
    data = request.json
    try:
        pat = upper(data.get('numero_patrimonio',''))
        e = Estabilizador.query.filter_by(numero_patrimonio=pat).first()
        if e:
            e.ativo = True
        else:
            e = Estabilizador(); db.session.add(e)
        e.numero_patrimonio   = pat
        e.pa                  = upper(data.get('pa'))
        e.marca               = upper(data.get('marca'))
        e.modelo              = upper(data.get('modelo'))
        e.potencia_va         = to_int(data.get('potencia_va'))
        e.localizacao         = upper(data.get('localizacao',''))
        e.empresa_id          = to_int(data.get('empresa_id'))
        e.data_aluguel        = _parse_date(data['data_aluguel'])
        e.data_fim_fidelidade = _parse_date(data['data_fim_fidelidade'])
        e.observacoes         = upper(data.get('observacoes'))
        db.session.commit()
        return jsonify({'sucesso': True, 'id': e.id}), 201
    except Exception as ex:
        db.session.rollback()
        return jsonify({'sucesso': False, 'erro': str(ex)[:150]}), 400

@app.route('/api/estabilizadores/<int:id>', methods=['PUT'])
@perfil_required('admin','tecnico')
def atualizar_estabilizador(id):
    e = db.get_or_404(Estabilizador, id); data = request.json
    try:
        registrar_movimentacao('Estabilizador', id, 'localizacao', e.localizacao, upper(data.get('localizacao', e.localizacao)))
        emp_ant = e.empresa.nome if e.empresa else ''
        nova_emp = Empresa.query.get(to_int(data.get('empresa_id', e.empresa_id)))
        registrar_movimentacao('Estabilizador', id, 'empresa', emp_ant, nova_emp.nome if nova_emp else '')
        e.numero_patrimonio=upper(data.get('numero_patrimonio',e.numero_patrimonio)); e.marca=upper(data.get('marca',e.marca)); e.pa=upper(data.get('pa',e.pa))
        pva2 = data.get('potencia_va', e.potencia_va)
        e.modelo=upper(data.get('modelo',e.modelo)); e.potencia_va=int(pva2) if pva2 not in (None,'','None') else None
        e.localizacao=upper(data.get('localizacao',e.localizacao)); e.empresa_id=data.get('empresa_id',e.empresa_id)
        e.observacoes=upper(data.get('observacoes',e.observacoes))
        if data.get('data_aluguel'):        e.data_aluguel=_parse_date(data['data_aluguel'])
        if data.get('data_fim_fidelidade'): e.data_fim_fidelidade=_parse_date(data['data_fim_fidelidade'])
        registrar_log('editar', 'Estabilizador', e.numero_patrimonio)
        db.session.commit(); return jsonify({'sucesso': True})
    except Exception as ex: db.session.rollback(); return jsonify({'sucesso': False, 'erro': str(ex)[:150]}), 400

@app.route('/api/estabilizadores/<int:id>', methods=['DELETE'])
@perfil_required('admin')
def deletar_estabilizador(id):
    e = db.get_or_404(Estabilizador, id); e.ativo=False; db.session.commit(); return jsonify({'sucesso': True})

# ── Notebooks ──────────────────────────────────────────────────────────────────
@app.route('/api/notebooks', methods=['POST'])
@perfil_required('admin','tecnico')
def criar_notebook():
    data = request.json
    try:
        pat = upper(data['numero_patrimonio'])
        n = Notebook.query.filter_by(numero_patrimonio=pat).first()
        if n:
            n.ativo = True
        else:
            n = Notebook(); db.session.add(n)
        n.numero_patrimonio   = pat
        n.pa                  = upper(data.get('pa'))
        n.marca               = upper(data.get('marca'))
        n.modelo              = upper(data.get('modelo'))
        n.processador         = upper(data.get('processador'))
        n.ram_gb              = to_int(data.get('ram_gb'))
        n.armazenamento_gb    = to_int(data.get('armazenamento_gb'))
        n.tamanho_tela        = to_int(data.get('tamanho_tela'))
        n.localizacao         = upper(data['localizacao'])
        n.empresa_id          = to_int(data['empresa_id'])
        n.data_aluguel        = _parse_date(data['data_aluguel'])
        n.data_fim_fidelidade = _parse_date(data['data_fim_fidelidade'])
        n.observacoes         = upper(data.get('observacoes'))
        db.session.commit()
        return jsonify({'sucesso': True, 'id': n.id}), 201
    except Exception as ex: db.session.rollback(); return jsonify({'sucesso': False, 'erro': str(ex)[:150]}), 400

@app.route('/api/notebooks/<int:id>', methods=['PUT'])
@perfil_required('admin','tecnico')
def atualizar_notebook(id):
    n = db.get_or_404(Notebook, id); data = request.json
    try:
        registrar_movimentacao('Notebook', id, 'localizacao', n.localizacao, upper(data.get('localizacao', n.localizacao)))
        emp_ant = n.empresa.nome if n.empresa else ''
        nova_emp = Empresa.query.get(to_int(data.get('empresa_id', n.empresa_id)))
        registrar_movimentacao('Notebook', id, 'empresa', emp_ant, nova_emp.nome if nova_emp else '')
        n.numero_patrimonio=upper(data.get('numero_patrimonio',n.numero_patrimonio)); n.marca=upper(data.get('marca',n.marca)); n.pa=upper(data.get('pa',n.pa))
        n.modelo=upper(data.get('modelo',n.modelo)); n.processador=upper(data.get('processador',n.processador))
        n.ram_gb=data.get('ram_gb',n.ram_gb); n.armazenamento_gb=data.get('armazenamento_gb',n.armazenamento_gb)
        n.tamanho_tela=data.get('tamanho_tela',n.tamanho_tela)
        n.localizacao=upper(data.get('localizacao',n.localizacao)); n.empresa_id=data.get('empresa_id',n.empresa_id)
        n.observacoes=upper(data.get('observacoes',n.observacoes))
        if data.get('data_aluguel'):        n.data_aluguel=_parse_date(data['data_aluguel'])
        if data.get('data_fim_fidelidade'): n.data_fim_fidelidade=_parse_date(data['data_fim_fidelidade'])
        registrar_log('editar', 'Notebook', n.numero_patrimonio)
        db.session.commit(); return jsonify({'sucesso': True})
    except Exception as ex: db.session.rollback(); return jsonify({'sucesso': False, 'erro': str(ex)[:150]}), 400

@app.route('/api/notebooks/<int:id>', methods=['DELETE'])
@perfil_required('admin')
def deletar_notebook(id):
    n = db.get_or_404(Notebook, id); n.ativo=False; db.session.commit(); return jsonify({'sucesso': True})

# ══════════════════════════════════════════════════════════════════════════════
# ROTAS — EMPRESAS
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/api/empresas', methods=['GET'])
@login_required
def listar_empresas():
    return jsonify([e.to_dict() for e in Empresa.query.order_by(Empresa.nome).all()])

@app.route('/api/empresas', methods=['POST'])
@perfil_required('admin','tecnico')
def criar_empresa():
    data = request.json
    try:
        e = Empresa(nome=upper(data['nome']), cnpj=data.get('cnpj'),
                    contato=upper(data.get('contato')), telefone=data.get('telefone'))
        db.session.add(e); db.session.commit()
        return jsonify({'sucesso': True, 'id': e.id}), 201
    except Exception as ex: db.session.rollback(); return jsonify({'sucesso': False, 'erro': 'Numero de patrimonio ja cadastrado. Use outro numero.' if 'UniqueViolation' in str(ex) or 'unicidade' in str(ex) else str(ex)[:150]}), 400

@app.route('/api/empresas/<int:id>', methods=['PUT'])
@perfil_required('admin','tecnico')
def atualizar_empresa(id):
    e = db.get_or_404(Empresa, id); data = request.json
    try:
        e.nome=upper(data.get('nome',e.nome)); e.cnpj=data.get('cnpj',e.cnpj)
        e.contato=upper(data.get('contato',e.contato)); e.telefone=data.get('telefone',e.telefone)
        db.session.commit(); return jsonify({'sucesso': True})
    except Exception as ex: db.session.rollback(); return jsonify({'sucesso': False, 'erro': 'Numero de patrimonio ja cadastrado. Use outro numero.' if 'UniqueViolation' in str(ex) or 'unicidade' in str(ex) else str(ex)[:150]}), 400

@app.route('/api/empresas/<int:id>', methods=['DELETE'])
@perfil_required('admin')
def deletar_empresa(id):
    e = db.get_or_404(Empresa, id); db.session.delete(e); db.session.commit(); return jsonify({'sucesso': True})

# ══════════════════════════════════════════════════════════════════════════════
# ROTAS — LOCALIZAÇÕES
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/api/localizacoes', methods=['GET'])
@login_required
def listar_localizacoes():
    return jsonify([l.to_dict() for l in Localizacao.query.order_by(Localizacao.nome).all()])

@app.route('/api/localizacoes', methods=['POST'])
@perfil_required('admin','tecnico')
def criar_localizacao():
    data = request.json
    try:
        l = Localizacao(nome=upper(data['nome']), descricao=upper(data.get('descricao')))
        db.session.add(l); db.session.commit()
        return jsonify({'sucesso': True, 'id': l.id}), 201
    except Exception as ex: db.session.rollback(); return jsonify({'sucesso': False, 'erro': 'Numero de patrimonio ja cadastrado. Use outro numero.' if 'UniqueViolation' in str(ex) or 'unicidade' in str(ex) else str(ex)[:150]}), 400

@app.route('/api/localizacoes/<int:id>', methods=['PUT'])
@perfil_required('admin','tecnico')
def atualizar_localizacao(id):
    l = db.get_or_404(Localizacao, id); data = request.json
    try:
        l.nome=upper(data.get('nome',l.nome)); l.descricao=upper(data.get('descricao',l.descricao))
        db.session.commit(); return jsonify({'sucesso': True})
    except Exception as ex: db.session.rollback(); return jsonify({'sucesso': False, 'erro': 'Numero de patrimonio ja cadastrado. Use outro numero.' if 'UniqueViolation' in str(ex) or 'unicidade' in str(ex) else str(ex)[:150]}), 400

@app.route('/api/localizacoes/<int:id>', methods=['DELETE'])
@perfil_required('admin')
def deletar_localizacao(id):
    l = db.get_or_404(Localizacao, id); db.session.delete(l); db.session.commit(); return jsonify({'sucesso': True})

# ══════════════════════════════════════════════════════════════════════════════
# ROTAS — RELATÓRIOS
# ══════════════════════════════════════════════════════════════════════════════


@app.route('/api/testar_smtp', methods=['POST'])
@perfil_required('admin')
def testar_smtp():
    """Envia e-mail de teste para validar configuração SMTP."""
    dest = request.json.get('email') or session.get('email', '')
    # Busca email do admin logado
    u = db.session.get(Usuario, session.get('usuario_id'))
    if not dest and u:
        dest = u.email
    if not dest:
        return jsonify({'sucesso': False, 'erro': 'Nenhum e-mail de destino'})
    ok, erro = enviar_email(
        dest,
        '🖥 PatrimônioTech — Teste de configuração SMTP',
        f"""<div style="font-family:Arial,sans-serif;padding:24px;max-width:480px;">
          <h2 style="color:#2D7DD2;">✅ SMTP configurado com sucesso!</h2>
          <p>Este é um e-mail de teste do sistema <strong>PatrimônioTech</strong>.</p>
          <p style="color:#666;font-size:13px;">Servidor: {os.getenv('SMTP_HOST')}:{os.getenv('SMTP_PORT')}<br>
          Usuário: {os.getenv('SMTP_USER')}</p>
        </div>"""
    )
    if ok:
        return jsonify({'sucesso': True, 'mensagem': f'E-mail de teste enviado para {dest}'})
    return jsonify({'sucesso': False, 'erro': erro})

@app.route('/relatorio/excel')
@login_required
def relatorio_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb   = Workbook()
    hfill = PatternFill("solid", fgColor="3d1236")
    hfont = Font(bold=True, color="FFFFFF", size=11)
    brd   = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'),  bottom=Side(style='thin')
    )
    red_fill = PatternFill("solid", fgColor="FFB3B3")

    def _header(ws, cols):
        for c, t in enumerate(cols, 1):
            cell = ws.cell(row=1, column=c, value=t)
            cell.font      = hfont
            cell.fill      = hfill
            cell.alignment = Alignment(horizontal='center')
            cell.border    = brd
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 20

    def _row(ws, r, values, alert_col):
        for c, v in enumerate(values, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = brd
            if c == alert_col and isinstance(v, date) and (v - date.today()).days <= 30:
                cell.fill = red_fill

    # ── Monitores ─────────────────────────────────────────────────────────
    ws1 = wb.active; ws1.title = "Monitores"
    _header(ws1, ['Patrimônio','PA','Marca','Modelo','Polegadas','Localização','Empresa','Aluguel','Fim Fidelidade','Obs'])
    for r, m in enumerate(Monitor.query.filter_by(ativo=True).all(), 2):
        _row(ws1, r, [m.numero_patrimonio, m.pa, m.marca, m.modelo, m.tamanho_polegadas,
             m.localizacao, m.empresa.nome if m.empresa else '',
             m.data_aluguel, m.data_fim_fidelidade, m.observacoes], alert_col=9)

    # ── Desktops ──────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Desktops")
    _header(ws2, ['Patrimônio','PA','Marca','Modelo','Processador','RAM (GB)','Storage (GB)','Localização','Empresa','Aluguel','Fim Fidelidade','Obs'])
    for r, d in enumerate(Desktop.query.filter_by(ativo=True).all(), 2):
        _row(ws2, r, [d.numero_patrimonio, d.pa, d.marca, d.modelo, d.processador,
             d.ram_gb, d.armazenamento_gb, d.localizacao,
             d.empresa.nome if d.empresa else '',
             d.data_aluguel, d.data_fim_fidelidade, d.observacoes], alert_col=11)

    # ── Notebooks ─────────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Notebooks")
    _header(ws3, ['Patrimônio','PA','Marca','Modelo','Processador','RAM (GB)','Storage (GB)','Tela (pol)','Localização','Empresa','Aluguel','Fim Fidelidade','Obs'])
    for r, n in enumerate(Notebook.query.filter_by(ativo=True).all(), 2):
        _row(ws3, r, [n.numero_patrimonio, n.pa, n.marca, n.modelo, n.processador,
             n.ram_gb, n.armazenamento_gb, n.tamanho_tela, n.localizacao,
             n.empresa.nome if n.empresa else '',
             n.data_aluguel, n.data_fim_fidelidade, n.observacoes], alert_col=12)

    # ── Estabilizadores ───────────────────────────────────────────────────
    ws4 = wb.create_sheet("Estabilizadores")
    _header(ws4, ['Patrimônio','PA','Marca','Modelo','Potência (VA)','Localização','Empresa','Aluguel','Fim Fidelidade','Obs'])
    for r, e in enumerate(Estabilizador.query.filter_by(ativo=True).all(), 2):
        _row(ws4, r, [e.numero_patrimonio, e.pa, e.marca, e.modelo, e.potencia_va,
             e.localizacao, e.empresa.nome if e.empresa else '',
             e.data_aluguel, e.data_fim_fidelidade, e.observacoes], alert_col=9)

    out = io.BytesIO()
    wb.save(out); out.seek(0)
    return send_file(out,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'patrimonio_{date.today()}.xlsx')

@app.route('/relatorio/pdf')
@login_required
def relatorio_pdf():
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    out = io.BytesIO()
    doc = SimpleDocTemplate(out, pagesize=landscape(A4),
                            leftMargin=20, rightMargin=20,
                            topMargin=20, bottomMargin=20)

    # ── Estilos de parágrafo ──────────────────────────────────────────────────
    # IMPORTANTE: TEXTCOLOR do TableStyle NÃO afeta Paragraphs — cor deve estar no ParagraphStyle
    cs_normal   = ParagraphStyle('cn',  fontSize=7.5, fontName='Helvetica',      leading=10, textColor=colors.HexColor('#1a0a18'))
    cs_header   = ParagraphStyle('ch',  fontSize=8,   fontName='Helvetica-Bold', leading=10, textColor=colors.white)
    cs_bold     = ParagraphStyle('cb',  fontSize=7.5, fontName='Helvetica-Bold', leading=10, textColor=colors.HexColor('#1a0a18'))
    cs_vencido  = ParagraphStyle('cv',  fontSize=7.5, fontName='Helvetica-Bold', leading=10, textColor=colors.HexColor('#dc2626'))
    cs_alerta   = ParagraphStyle('ca',  fontSize=7.5, fontName='Helvetica-Bold', leading=10, textColor=colors.HexColor('#d97706'))

    def P(text, style='normal'):
        estilos = {'normal': cs_normal, 'header': cs_header, 'bold': cs_bold,
                   'vencido': cs_vencido, 'alerta': cs_alerta}
        return Paragraph(str(text or ''), estilos.get(style, cs_normal))

    base_ts = [
        ('BACKGROUND',    (0,0), (-1,0),  colors.HexColor('#3d1236')),
        ('ROWBACKGROUNDS',(0,1), (-1,-1), [colors.white, colors.HexColor('#F7EDF5')]),
        ('GRID',          (0,0), (-1,-1), 0.4, colors.HexColor('#E8D8E5')),
        ('ALIGN',         (0,0), (-1,-1), 'LEFT'),
        ('VALIGN',        (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING',    (0,0), (-1,-1), 4),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ('LEFTPADDING',   (0,0), (-1,-1), 5),
        ('RIGHTPADDING',  (0,0), (-1,-1), 5),
    ]

    hs = ParagraphStyle('h', fontSize=11, fontName='Helvetica-Bold',
                        textColor=colors.HexColor('#7b2d6e'),
                        spaceBefore=10, spaceAfter=4)

    elems = []
    elems.append(Paragraph("Relatório de Controle de Patrimônios",
        ParagraphStyle('t', fontSize=15, fontName='Helvetica-Bold',
                       textColor=colors.HexColor('#3d1236'), spaceAfter=4)))
    elems.append(Paragraph(f"Gerado em {date.today().strftime('%d/%m/%Y')}",
        ParagraphStyle('s', fontSize=9, fontName='Helvetica',
                       textColor=colors.grey, spaceAfter=12)))

    FILL_VENCIDO = colors.HexColor('#fce8e8')  # fundo vermelho claro
    FILL_ALERTA  = colors.HexColor('#fef3cd')  # fundo laranja claro

    def _add(title, rows_raw, widths):
        rows_fmt  = []
        status_map = {}
        for r_idx, row in enumerate(rows_raw):
            if r_idx == 0:
                # Cabeçalho — texto branco sobre fundo roxo
                rows_fmt.append([P(c, 'header') for c in row])
            else:
                last_val = row[-1]
                txt, status = last_val if isinstance(last_val, tuple) else (str(last_val), 'ok')
                # Aplica estilo de cor correto na célula de data
                fim_cell = P(txt, status if status != 'ok' else 'normal')
                # Demais células ficam normais mas ficam bold se a linha for crítica
                other_style = 'bold' if status != 'ok' else 'normal'
                fmt_row = [P(c.text if isinstance(c, Paragraph) else str(c or ''), other_style)
                           for c in row[:-1]] + [fim_cell]
                rows_fmt.append(fmt_row)
                if status != 'ok':
                    status_map[r_idx] = status

        elems.append(Paragraph(title, hs))
        t  = Table(rows_fmt, colWidths=widths, repeatRows=1)
        ts = TableStyle(base_ts[:])

        # Fundo colorido nas linhas críticas (BACKGROUND funciona mesmo com Paragraph)
        for r_idx, status in status_map.items():
            fill = FILL_VENCIDO if status == 'vencido' else FILL_ALERTA
            ts.add('BACKGROUND', (0, r_idx), (-1, r_idx), fill)

        t.setStyle(ts)
        elems.append(t)
        elems.append(Spacer(1, 12))

    def _fim(obj):
        # Retorna (texto, status) onde status: 'vencido','alerta','ok'
        if not obj.data_fim_fidelidade:
            return ('', 'ok')
        dias = obj.dias_para_vencer()
        txt  = obj.data_fim_fidelidade.strftime('%d/%m/%Y')
        if dias is not None and dias < 0:
            return (txt + ' ⚠ VENCIDO', 'vencido')
        if dias is not None and dias <= 30:
            return (txt + ' ⚠', 'alerta')
        return (txt, 'ok')

    # Larguras somam 801pt (landscape A4 - margens)
    # Monitores: Patrimônio|PA|Marca/Modelo|Localização|Empresa|Aluguel|Fim Fid.
    rows_m = [['Patrimônio','PA','Marca/Modelo','Localização','Empresa','Aluguel','Fim Fid.']]
    for m in Monitor.query.filter_by(ativo=True).all():
        rows_m.append([P(m.numero_patrimonio, 'bold'), P(m.pa),
            P(f"{m.marca or ''} {m.modelo or ''}".strip()),
            P(m.localizacao),
            P(m.empresa.nome if m.empresa else ''),
            P(m.data_aluguel.strftime('%d/%m/%Y') if m.data_aluguel else ''),
            _fim(m)])
    _add("▪ Monitores", rows_m, [95, 60, 130, 165, 175, 80, 96])

    # Desktops: Patrimônio|PA|Marca/Modelo|Processador|Localização|Empresa|Aluguel|Fim Fid.
    rows_d = [['Patrimônio','PA','Marca/Modelo','Processador','Localização','Empresa','Aluguel','Fim Fid.']]
    for d in Desktop.query.filter_by(ativo=True).all():
        rows_d.append([P(d.numero_patrimonio, 'bold'), P(d.pa),
            P(f"{d.marca or ''} {d.modelo or ''}".strip()),
            P(d.processador),
            P(d.localizacao),
            P(d.empresa.nome if d.empresa else ''),
            P(d.data_aluguel.strftime('%d/%m/%Y') if d.data_aluguel else ''),
            _fim(d)])
    _add("▪ Desktops", rows_d, [90, 50, 110, 110, 130, 150, 75, 86])

    # Estabilizadores: Patrimônio|PA|Marca/Modelo|Potência|Localização|Empresa|Aluguel|Fim Fid.
    rows_e = [['Patrimônio','PA','Marca/Modelo','Potência VA','Localização','Empresa','Aluguel','Fim Fid.']]
    for e in Estabilizador.query.filter_by(ativo=True).all():
        rows_e.append([P(e.numero_patrimonio, 'bold'), P(e.pa),
            P(f"{e.marca or ''} {e.modelo or ''}".strip()),
            P(str(e.potencia_va or '')),
            P(e.localizacao),
            P(e.empresa.nome if e.empresa else ''),
            P(e.data_aluguel.strftime('%d/%m/%Y') if e.data_aluguel else ''),
            _fim(e)])
    _add("▪ Estabilizadores", rows_e, [90, 50, 110, 65, 130, 175, 75, 106])

    # Notebooks: Patrimônio|PA|Marca/Modelo|Processador|Tela|Localização|Empresa|Aluguel|Fim Fid.
    rows_n = [['Patrimônio','PA','Marca/Modelo','Processador','Tela','Localização','Empresa','Aluguel','Fim Fid.']]
    for n in Notebook.query.filter_by(ativo=True).all():
        rows_n.append([P(n.numero_patrimonio, 'bold'), P(n.pa),
            P(f"{n.marca or ''} {n.modelo or ''}".strip()),
            P(n.processador),
            P(str(n.tamanho_tela or '')),
            P(n.localizacao),
            P(n.empresa.nome if n.empresa else ''),
            P(n.data_aluguel.strftime('%d/%m/%Y') if n.data_aluguel else ''),
            _fim(n)])
    _add("▪ Notebooks", rows_n, [88, 48, 105, 98, 38, 115, 150, 72, 87])

    doc.build(elems)
    out.seek(0)
    return send_file(out, mimetype='application/pdf', as_attachment=False,
                     download_name=f'patrimonio_{date.today()}.pdf')


# ══════════════════════════════════════════════════════════════════════════════
# ROTAS — CHAMADOS
# ══════════════════════════════════════════════════════════════════════════════

def _get_equipamento(tipo, eid):
    # Normaliza: aceita 'monitor' ou 'Monitor'
    tipo_norm = tipo.capitalize() if tipo else tipo
    m = {
        'Monitor':       Monitor,
        'Desktop':       Desktop,
        'Estabilizador': Estabilizador,
        'Notebook':      Notebook,
    }
    model = m.get(tipo_norm)
    if not model: return None
    return model.query.get(eid)

@app.route('/api/chamados', methods=['GET'])
@login_required
def listar_chamados():
    tipo = request.args.get('tipo', '')
    eid  = request.args.get('equipamento_id', '')
    q    = Chamado.query
    if tipo: q = q.filter_by(tipo_equipamento=tipo)
    if eid:  q = q.filter_by(equipamento_id=int(eid))
    chamados = q.order_by(Chamado.data_abertura.desc()).all()
    result = []
    for c in chamados:
        d = c.to_dict()
        eq = _get_equipamento(c.tipo_equipamento, c.equipamento_id)
        d['patrimonio']  = eq.numero_patrimonio if eq else '—'
        d['pa']          = eq.pa if eq else '—'
        d['empresa']     = eq.empresa.nome if eq and eq.empresa else '—'
        d['localizacao'] = eq.localizacao if eq else '—'
        result.append(d)
    return jsonify(result)

@app.route('/api/chamados/stats', methods=['GET'])
@login_required
def stats_chamados():
    total    = Chamado.query.count()
    abertos  = Chamado.query.filter_by(status='aberto').count()
    fechados = Chamado.query.filter_by(status='fechado').count()
    from sqlalchemy import func
    top = (db.session.query(Chamado.tipo_equipamento, Chamado.equipamento_id,
                            func.count(Chamado.id).label('total'))
           .group_by(Chamado.tipo_equipamento, Chamado.equipamento_id)
           .order_by(func.count(Chamado.id).desc())
           .limit(5).all())
    top_list = []
    for tipo, eid, cnt in top:
        eq = _get_equipamento(tipo, eid)
        top_list.append({'tipo': tipo,
                         'patrimonio': eq.numero_patrimonio if eq else '—',
                         'pa': eq.pa if eq else '—',
                         'total': cnt})
    # Stats por tipo
    por_tipo = {}
    for tipo in ['Monitor','Desktop','Estabilizador','Notebook']:
        por_tipo[tipo] = Chamado.query.filter_by(tipo_equipamento=tipo).count()
    return jsonify({'total': total, 'abertos': abertos, 'fechados': fechados,
                    'top': top_list, 'por_tipo': por_tipo})

@app.route('/api/chamados', methods=['POST'])
@perfil_required('admin', 'tecnico')
def criar_chamado():
    data = request.json
    try:
        c = Chamado(
            numero_chamado   = data['numero_chamado'].strip().upper(),
            tipo_equipamento = data['tipo_equipamento'],
            equipamento_id   = int(data['equipamento_id']),
            data_abertura    = _parse_date(data['data_abertura']),
            data_solucao     = _parse_date(data['data_solucao']) if data.get('data_solucao') else None,
            descricao        = (data.get('descricao') or '').strip().upper(),
            solucao          = (data.get('solucao') or '').strip().upper(),
            status           = data.get('status', 'aberto'),
            prioridade       = data.get('prioridade', 'media'),
            status_detalhe   = data.get('status_detalhe', 'em_analise'),
        )
        db.session.add(c); db.session.commit()
        return jsonify({'sucesso': True, 'id': c.id}), 201
    except Exception as ex:
        db.session.rollback()
        return jsonify({'sucesso': False, 'erro': str(ex)[:150]}), 400

@app.route('/api/chamados/<int:id>', methods=['PUT'])
@perfil_required('admin', 'tecnico')
def atualizar_chamado(id):
    c = db.get_or_404(Chamado, id)
    data = request.json
    try:
        if data.get('numero_chamado'): c.numero_chamado   = data['numero_chamado'].strip().upper()
        if data.get('tipo_equipamento'): c.tipo_equipamento = data['tipo_equipamento']
        if data.get('equipamento_id'):   c.equipamento_id   = int(data['equipamento_id'])
        if data.get('data_abertura'):    c.data_abertura    = _parse_date(data['data_abertura'])
        c.data_solucao = _parse_date(data['data_solucao']) if data.get('data_solucao') else None
        c.descricao    = (data.get('descricao') or '').strip().upper()
        c.solucao      = (data.get('solucao') or '').strip().upper()
        if data.get('status'): c.status = data['status']
        if data.get('prioridade'):     c.prioridade     = data['prioridade']
        if data.get('status_detalhe'): c.status_detalhe = data['status_detalhe']
        db.session.commit()
        return jsonify({'sucesso': True})
    except Exception as ex:
        db.session.rollback()
        return jsonify({'sucesso': False, 'erro': str(ex)[:150]}), 400

@app.route('/api/chamados/<int:id>', methods=['DELETE'])
@perfil_required('admin')
def deletar_chamado(id):
    c = db.get_or_404(Chamado, id)
    db.session.delete(c); db.session.commit()
    return jsonify({'sucesso': True})

@app.route('/relatorio/chamados/excel')
@login_required
def relatorio_chamados_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    wb   = Workbook()
    ws   = wb.active; ws.title = "Chamados"
    hfill = PatternFill("solid", fgColor="3d1236")
    hfont = Font(bold=True, color="FFFFFF", size=11)
    brd   = Border(left=Side(style='thin'), right=Side(style='thin'),
                   top=Side(style='thin'), bottom=Side(style='thin'))
    cols  = ['Nº Chamado','Tipo','Patrimônio','PA','Empresa','Localização',
             'Data Abertura','Data Solução','Status','Descrição','Solução']
    for c, t in enumerate(cols, 1):
        cell = ws.cell(row=1, column=c, value=t)
        cell.font = hfont; cell.fill = hfill
        cell.alignment = Alignment(horizontal='center'); cell.border = brd
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 20
    red_fill    = PatternFill("solid", fgColor="FFB3B3")
    green_fill  = PatternFill("solid", fgColor="B3FFB3")
    for r, ch in enumerate(Chamado.query.order_by(Chamado.data_abertura.desc()).all(), 2):
        eq = _get_equipamento(ch.tipo_equipamento, ch.equipamento_id)
        row = [ch.numero_chamado, ch.tipo_equipamento,
               eq.numero_patrimonio if eq else '—',
               eq.pa if eq else '—',
               eq.empresa.nome if eq and eq.empresa else '—',
               eq.localizacao if eq else '—',
               ch.data_abertura.strftime('%d/%m/%Y') if ch.data_abertura else '',
               ch.data_solucao.strftime('%d/%m/%Y') if ch.data_solucao else '',
               ch.status.upper(), ch.descricao or '', ch.solucao or '']
        for c, v in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=v); cell.border = brd
            if ch.status == 'aberto':   cell.fill = red_fill
            if ch.status == 'fechado':  cell.fill = green_fill
    out = io.BytesIO(); wb.save(out); out.seek(0)
    return send_file(out,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True, download_name=f'chamados_{date.today()}.xlsx')



@app.route('/api/busca')
@login_required
def busca_global():
    q = request.args.get('q', '').strip().upper()
    if len(q) < 2:
        return jsonify({'equipamentos': [], 'chamados': [], 'localizacoes': []})
    pat = f'%{q}%'
    equips = []
    for Model, tipo in [(Monitor,'Monitor'),(Desktop,'Desktop'),
                        (Estabilizador,'Estabilizador'),(Notebook,'Notebook')]:
        for eq in Model.query.filter(
            Model.ativo==True,
            db.or_(Model.numero_patrimonio.ilike(pat),
                   Model.pa.ilike(pat),
                   Model.marca.ilike(pat),
                   Model.modelo.ilike(pat),
                   Model.localizacao.ilike(pat))
        ).limit(5).all():
            d = eq.to_dict()
            equips.append(d)
    chamados = [c.to_dict() | {
        'patrimonio': (lambda eq: eq.numero_patrimonio if eq else '—')(_get_equipamento(c.tipo_equipamento, c.equipamento_id))
    } for c in Chamado.query.filter(
        db.or_(Chamado.numero_chamado.ilike(pat),
               Chamado.descricao.ilike(pat))
    ).limit(5).all()]
    locs = [l.to_dict() for l in Localizacao.query.filter(
        Localizacao.nome.ilike(pat)
    ).limit(5).all()]
    return jsonify({'equipamentos': equips, 'chamados': chamados, 'localizacoes': locs})

@app.route('/api/stats/graficos')
@login_required
def stats_graficos():
    from sqlalchemy import func
    # Equipamentos por tipo
    por_tipo = {
        'Monitor':       Monitor.query.filter_by(ativo=True).count(),
        'Desktop':       Desktop.query.filter_by(ativo=True).count(),
        'Estabilizador': Estabilizador.query.filter_by(ativo=True).count(),
        'Notebook':      Notebook.query.filter_by(ativo=True).count(),
    }
    # Chamados por prioridade
    prioridades = {}
    for p in ['baixa','media','alta','critica']:
        prioridades[p] = Chamado.query.filter_by(prioridade=p).count()
    # Chamados por status detalhe
    status_det = {}
    for s in ['em_analise','aguardando_peca','em_reparo','resolvido']:
        status_det[s] = Chamado.query.filter_by(status_detalhe=s).count()
    # Chamados por mes (ultimos 6 meses)
    meses = []
    hoje = date.today()
    for i in range(5, -1, -1):
        m = hoje.month - i
        y = hoje.year
        while m <= 0: m += 12; y -= 1
        inicio = date(y, m, 1)
        import calendar
        fim = date(y, m, calendar.monthrange(y, m)[1])
        count = Chamado.query.filter(
            Chamado.data_abertura >= inicio,
            Chamado.data_abertura <= fim
        ).count()
        meses.append({'mes': inicio.strftime('%b/%y'), 'total': count})
    # Equipamentos vencidos / a vencer / ok
    limite = date.today() + timedelta(days=30)
    vencidos = av = ok = 0
    for Model in [Monitor, Desktop, Estabilizador, Notebook]:
        for eq in Model.query.filter_by(ativo=True).all():
            if not eq.data_fim_fidelidade: continue
            if eq.data_fim_fidelidade < date.today(): vencidos += 1
            elif eq.data_fim_fidelidade <= limite: av += 1
            else: ok += 1
    return jsonify({
        'por_tipo': por_tipo,
        'prioridades': prioridades,
        'status_detalhe': status_det,
        'chamados_por_mes': meses,
        'contratos': {'vencidos': vencidos, 'a_vencer': av, 'ok': ok},
    })


@app.route('/api/importar', methods=['POST'])
@perfil_required('admin','tecnico')
def importar_equipamentos():
    import csv, io as _io
    if 'arquivo' not in request.files:
        return jsonify({'sucesso': False, 'erro': 'Nenhum arquivo enviado'}), 400
    f   = request.files['arquivo']
    ext = f.filename.rsplit('.', 1)[-1].lower()
    rows = []
    try:
        if ext == 'csv':
            content = f.read().decode('utf-8-sig')
            reader = csv.DictReader(_io.StringIO(content))
            rows = list(reader)
        elif ext in ('xlsx','xls'):
            from openpyxl import load_workbook
            wb = load_workbook(f, read_only=True, data_only=True)
            ws = wb.active
            headers = [str(c.value or '').strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append({headers[i]: (str(v).strip() if v is not None else '') for i, v in enumerate(row)})
        else:
            return jsonify({'sucesso': False, 'erro': 'Formato não suportado. Use CSV ou XLSX'}), 400
    except Exception as ex:
        return jsonify({'sucesso': False, 'erro': f'Erro ao ler arquivo: {str(ex)[:150]}'}), 400

    ok = erros = 0
    detalhes = []
    for i, row in enumerate(rows, 2):
        try:
            tipo = upper(row.get('tipo',''))
            pat  = upper(row.get('numero_patrimonio', row.get('patrimonio','')))
            if not tipo or not pat:
                detalhes.append(f'Linha {i}: tipo ou patrimônio vazio')
                erros += 1; continue
            emp_nome = upper(row.get('empresa',''))
            emp = Empresa.query.filter(db.func.upper(Empresa.nome)==emp_nome).first() if emp_nome else None
            kwargs = dict(
                numero_patrimonio   = pat,
                pa                  = upper(row.get('pa')),
                marca               = upper(row.get('marca')),
                modelo              = upper(row.get('modelo')),
                localizacao         = upper(row.get('localizacao','')),
                empresa_id          = emp.id if emp else None,
                data_aluguel        = _parse_date(row.get('data_aluguel','')) if row.get('data_aluguel') else None,
                data_fim_fidelidade = _parse_date(row.get('data_fim_fidelidade','')) if row.get('data_fim_fidelidade') else None,
                observacoes         = upper(row.get('observacoes')),
                ativo               = True,
            )
            ModelMap = {'MONITOR': Monitor,'DESKTOP': Desktop,'ESTABILIZADOR': Estabilizador,'NOTEBOOK': Notebook}
            Model = ModelMap.get(tipo)
            if not Model:
                detalhes.append(f'Linha {i}: tipo "{tipo}" inválido')
                erros += 1; continue
            existing = Model.query.filter_by(numero_patrimonio=pat).first()
            if existing:
                for k,v in kwargs.items(): setattr(existing, k, v)
            else:
                obj = Model(**kwargs)
                # Campos extras por tipo
                if tipo == 'MONITOR':
                    obj.tamanho_polegadas = row.get('tamanho_polegadas') or None
                elif tipo in ('DESKTOP','NOTEBOOK'):
                    obj.processador = upper(row.get('processador'))
                    obj.ram_gb = to_int(row.get('ram_gb'))
                    obj.armazenamento_gb = to_int(row.get('armazenamento_gb'))
                    if tipo == 'NOTEBOOK': obj.tamanho_tela = to_int(row.get('tamanho_tela'))
                elif tipo == 'ESTABILIZADOR':
                    obj.potencia_va = to_int(row.get('potencia_va'))
                db.session.add(obj)
            ok += 1
        except Exception as ex:
            detalhes.append(f'Linha {i}: {str(ex)[:100]}')
            erros += 1
    try:
        db.session.commit()
    except Exception as ex:
        db.session.rollback()
        return jsonify({'sucesso': False, 'erro': str(ex)[:200]}), 400
    return jsonify({'sucesso': True, 'importados': ok, 'erros': erros, 'detalhes': detalhes})

# ══════════════════════════════════════════════════════════════════════════════
# ROTAS — MOVIMENTAÇÕES
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/api/movimentacoes')
@login_required
def listar_movimentacoes():
    tipo = request.args.get('tipo', '')
    eid  = request.args.get('equipamento_id', '')
    q    = Movimentacao.query
    if tipo: q = q.filter_by(tipo_equipamento=tipo)
    if eid:  q = q.filter_by(equipamento_id=int(eid))
    movs = q.order_by(Movimentacao.criado_em.desc()).limit(200).all()
    result = []
    for mv in movs:
        d = mv.to_dict()
        eq = _get_equipamento(mv.tipo_equipamento, mv.equipamento_id)
        d['patrimonio'] = eq.numero_patrimonio if eq else '—'
        result.append(d)
    return jsonify(result)

# ══════════════════════════════════════════════════════════════════════════════
# ROTAS — LOG DE ATIVIDADES
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/api/logs')
@perfil_required('admin')
def listar_logs():
    page  = int(request.args.get('page', 1))
    limit = 50
    logs  = (LogAtividade.query
             .order_by(LogAtividade.criado_em.desc())
             .offset((page-1)*limit).limit(limit).all())
    total = LogAtividade.query.count()
    return jsonify({'logs': [l.to_dict() for l in logs], 'total': total, 'page': page})

# ══════════════════════════════════════════════════════════════════════════════
# ROTAS — COMENTÁRIOS DE CHAMADOS
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/api/chamados/<int:id>/comentarios', methods=['GET'])
@login_required
def listar_comentarios(id):
    coms = (ComentarioChamado.query
            .filter_by(chamado_id=id)
            .order_by(ComentarioChamado.criado_em.asc()).all())
    return jsonify([c.to_dict() for c in coms])

@app.route('/api/chamados/<int:id>/comentarios', methods=['POST'])
@login_required
def criar_comentario(id):
    data = request.json
    texto = (data.get('texto') or '').strip()
    if not texto:
        return jsonify({'sucesso': False, 'erro': 'Texto obrigatório'}), 400
    c = ComentarioChamado(chamado_id=id, usuario_id=session.get('usuario_id'), texto=texto)
    db.session.add(c); db.session.commit()
    registrar_log('comentar', 'Chamado', f'Chamado #{id}')
    return jsonify({'sucesso': True, 'comentario': c.to_dict()}), 201

@app.route('/api/chamados/<int:cid>/comentarios/<int:comid>', methods=['DELETE'])
@perfil_required('admin')
def deletar_comentario(cid, comid):
    c = db.get_or_404(ComentarioChamado, comid)
    db.session.delete(c); db.session.commit()
    return jsonify({'sucesso': True})

# ══════════════════════════════════════════════════════════════════════════════
# ROTAS — RELATÓRIOS DE CHAMADOS
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/relatorio/chamados/pdf')
@login_required
def relatorio_chamados_pdf():
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    data_ini = request.args.get('data_ini', '')
    data_fim = request.args.get('data_fim', '')
    status   = request.args.get('status', '')

    tipo_eq  = request.args.get('tipo', '')
    q = Chamado.query
    if data_ini:
        try: q = q.filter(Chamado.data_abertura >= _parse_date(data_ini))
        except: pass
    if data_fim:
        try: q = q.filter(Chamado.data_abertura <= _parse_date(data_fim))
        except: pass
    if status:
        q = q.filter_by(status=status)
    if tipo_eq:
        q = q.filter_by(tipo_equipamento=tipo_eq)
    chamados = q.order_by(Chamado.data_abertura.desc()).all()

    out = io.BytesIO()
    doc = SimpleDocTemplate(out, pagesize=landscape(A4),
                            leftMargin=20, rightMargin=20, topMargin=20, bottomMargin=20)
    cs  = ParagraphStyle('c', fontSize=7.5, fontName='Helvetica', leading=10)
    csb = ParagraphStyle('cb', fontSize=7.5, fontName='Helvetica-Bold', leading=10)
    def P(t, b=False): return Paragraph(str(t or ''), csb if b else cs)
    base_ts = [
        ('BACKGROUND',    (0,0),(-1,0),  colors.HexColor('#3d1236')),
        ('TEXTCOLOR',     (0,0),(-1,0),  colors.white),
        ('FONTNAME',      (0,0),(-1,0),  'Helvetica-Bold'),
        ('FONTSIZE',      (0,0),(-1,0),  8),
        ('ROWBACKGROUNDS',(0,1),(-1,-1), [colors.white, colors.HexColor('#F7EDF5')]),
        ('GRID',          (0,0),(-1,-1), 0.4, colors.HexColor('#E8D8E5')),
        ('ALIGN',         (0,0),(-1,-1), 'LEFT'),
        ('VALIGN',        (0,0),(-1,-1), 'MIDDLE'),
        ('TOPPADDING',    (0,0),(-1,-1), 4),
        ('BOTTOMPADDING', (0,0),(-1,-1), 4),
        ('LEFTPADDING',   (0,0),(-1,-1), 5),
        ('RIGHTPADDING',  (0,0),(-1,-1), 5),
    ]
    elems = []
    titulo = "Relatório de Chamados de Manutenção"
    periodo = ""
    if data_ini or data_fim:
        periodo = f"  |  Período: {data_ini or '...'} até {data_fim or '...'}"
    if status:
        periodo += f"  |  Status: {status.upper()}"
    elems.append(Paragraph(titulo, ParagraphStyle('t', fontSize=14, fontName='Helvetica-Bold',
                           textColor=colors.HexColor('#3d1236'), spaceAfter=4)))
    elems.append(Paragraph(f"Gerado em {date.today().strftime('%d/%m/%Y')}{periodo}",
                           ParagraphStyle('s', fontSize=9, fontName='Helvetica',
                           textColor=colors.grey, spaceAfter=12)))
    # Resumo
    abertos  = sum(1 for c in chamados if c.status == 'aberto')
    fechados = sum(1 for c in chamados if c.status == 'fechado')
    elems.append(Paragraph(
        f"Total: {len(chamados)} chamados  |  Abertos: {abertos}  |  Fechados: {fechados}",
        ParagraphStyle('r', fontSize=10, fontName='Helvetica-Bold',
                       textColor=colors.HexColor('#7b2d6e'), spaceAfter=10)))

    rows = [['Nº Chamado','Tipo','Patrimônio','PA','Empresa','Localização',
             'Abertura','Solução','Status','Descrição']]
    for ch in chamados:
        eq = _get_equipamento(ch.tipo_equipamento, ch.equipamento_id)
        rows.append([
            P(ch.numero_chamado, True),
            P(ch.tipo_equipamento),
            P(eq.numero_patrimonio if eq else '—'),
            P(eq.pa if eq else '—'),
            P(eq.empresa.nome if eq and eq.empresa else '—'),
            P(eq.localizacao if eq else '—'),
            P(ch.data_abertura.strftime('%d/%m/%Y') if ch.data_abertura else ''),
            P(ch.data_solucao.strftime('%d/%m/%Y') if ch.data_solucao else '—'),
            P(ch.status.upper(), True),
            P(ch.descricao),
        ])
    t  = Table(rows, colWidths=[90,75,80,55,120,100,68,68,55,90], repeatRows=1)
    ts = TableStyle(base_ts[:])
    for i, ch in enumerate(chamados, 1):
        if ch.status == 'aberto':
            ts.add('BACKGROUND', (8,i),(8,i), colors.HexColor('#fef2f2'))
        else:
            ts.add('BACKGROUND', (8,i),(8,i), colors.HexColor('#f0fdf4'))
    t.setStyle(ts)
    elems.append(t)
    doc.build(elems)
    out.seek(0)
    return send_file(out, mimetype='application/pdf', as_attachment=False,
                     download_name=f'chamados_{date.today()}.pdf')

# ══════════════════════════════════════════════════════════════════════════════
# ROTA — FICHA DO EQUIPAMENTO (PDF)
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/relatorio/ficha/<tipo>/<int:id>')
@login_required
def ficha_equipamento(tipo, id):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable

    # Normaliza tipo: 'monitor' -> 'Monitor', 'estabilizador' -> 'Estabilizador'
    tipo_map = {'monitor':'Monitor','desktop':'Desktop',
                'estabilizador':'Estabilizador','notebook':'Notebook'}
    tipo = tipo_map.get(tipo.lower(), tipo.capitalize())

    eq = _get_equipamento(tipo, id)
    if not eq:
        return jsonify({'erro': 'Equipamento não encontrado'}), 404

    chamados = (Chamado.query.filter_by(tipo_equipamento=tipo, equipamento_id=id)
                .order_by(Chamado.data_abertura.desc()).all())
    movs     = (Movimentacao.query.filter_by(tipo_equipamento=tipo, equipamento_id=id)
                .order_by(Movimentacao.criado_em.desc()).limit(20).all())

    out = io.BytesIO()
    doc = SimpleDocTemplate(out, pagesize=A4,
                            leftMargin=40, rightMargin=40, topMargin=40, bottomMargin=40)
    roxo   = colors.HexColor('#3d1236')
    roxoL  = colors.HexColor('#7b2d6e')
    borda  = colors.HexColor('#e8d8e5')
    lilasB = colors.HexColor('#f7edf5')
    cs  = ParagraphStyle('c', fontSize=9, fontName='Helvetica', leading=13)
    csb = ParagraphStyle('cb', fontSize=9, fontName='Helvetica-Bold', leading=13)
    def P(t, b=False): return Paragraph(str(t or '—'), csb if b else cs)
    def campo(label, valor):
        return [[P(label, True), P(valor)]]

    elems = []
    # Cabeçalho
    elems.append(Paragraph("PatrimônioTech — Fama Soluções",
        ParagraphStyle('brand', fontSize=10, fontName='Helvetica', textColor=roxoL, spaceAfter=2)))
    elems.append(Paragraph(f"Ficha de Equipamento — {tipo}",
        ParagraphStyle('title', fontSize=18, fontName='Helvetica-Bold', textColor=roxo, spaceAfter=4)))
    elems.append(Paragraph(f"Gerada em {date.today().strftime('%d/%m/%Y')}",
        ParagraphStyle('sub', fontSize=9, fontName='Helvetica', textColor=colors.grey, spaceAfter=16)))
    elems.append(HRFlowable(width="100%", thickness=2, color=roxo, spaceAfter=16))

    # Dados principais
    ts_base = [
        ('BACKGROUND', (0,0),(0,-1), lilasB),
        ('FONTNAME',   (0,0),(0,-1), 'Helvetica-Bold'),
        ('FONTSIZE',   (0,0),(-1,-1), 9),
        ('GRID',       (0,0),(-1,-1), 0.5, borda),
        ('TOPPADDING', (0,0),(-1,-1), 6),
        ('BOTTOMPADDING',(0,0),(-1,-1), 6),
        ('LEFTPADDING',(0,0),(-1,-1), 8),
        ('RIGHTPADDING',(0,0),(-1,-1), 8),
        ('VALIGN',     (0,0),(-1,-1), 'MIDDLE'),
    ]
    elems.append(Paragraph("Identificação",
        ParagraphStyle('h2', fontSize=11, fontName='Helvetica-Bold', textColor=roxoL, spaceBefore=8, spaceAfter=6)))
    dados_id = [
        [P('Nº Patrimônio', True), P(eq.numero_patrimonio, True),
         P('PA', True), P(eq.pa)],
        [P('Marca', True), P(eq.marca),
         P('Modelo', True), P(eq.modelo)],
        [P('Localização', True), P(eq.localizacao),
         P('Empresa', True), P(eq.empresa.nome if eq.empresa else '—')],
        [P('Data Aluguel', True), P(eq.data_aluguel.strftime('%d/%m/%Y') if eq.data_aluguel else '—'),
         P('Fim Fidelidade', True), P(eq.data_fim_fidelidade.strftime('%d/%m/%Y') if eq.data_fim_fidelidade else '—')],
    ]
    if hasattr(eq, 'processador') and eq.processador:
        dados_id.append([P('Processador', True), P(eq.processador),
                         P('RAM', True), P(f"{eq.ram_gb or '—'} GB")])
    if hasattr(eq, 'tamanho_polegadas') and eq.tamanho_polegadas:
        dados_id.append([P('Tamanho', True), P(str(eq.tamanho_polegadas) + '"'), P('',''), P('')])
    if hasattr(eq, 'potencia_va') and eq.potencia_va:
        dados_id.append([P('Potência VA', True), P(str(eq.potencia_va)), P('',''), P('')])
    if eq.observacoes:
        dados_id.append([P('Observações', True), Paragraph(eq.observacoes or '', cs), P('',''), P('')])

    t_id = Table(dados_id, colWidths=[90,160,90,160])
    t_id.setStyle(TableStyle(ts_base))
    elems.append(t_id)
    elems.append(Spacer(1, 16))

    # Chamados
    elems.append(Paragraph(f"Chamados de Manutenção ({len(chamados)} total)",
        ParagraphStyle('h2', fontSize=11, fontName='Helvetica-Bold', textColor=roxoL, spaceBefore=8, spaceAfter=6)))
    if chamados:
        rows_ch = [['Nº Chamado','Abertura','Solução','Status','Descrição']]
        for ch in chamados:
            rows_ch.append([
                P(ch.numero_chamado, True),
                P(ch.data_abertura.strftime('%d/%m/%Y') if ch.data_abertura else ''),
                P(ch.data_solucao.strftime('%d/%m/%Y') if ch.data_solucao else '—'),
                P(ch.status.upper()),
                Paragraph(ch.descricao or '—', cs),
            ])
        t_ch = Table(rows_ch, colWidths=[100,70,70,60,200], repeatRows=1)
        ts_ch = TableStyle([
            ('BACKGROUND',    (0,0),(-1,0), roxo),
            ('TEXTCOLOR',     (0,0),(-1,0), colors.white),
            ('FONTNAME',      (0,0),(-1,0), 'Helvetica-Bold'),
            ('FONTSIZE',      (0,0),(-1,-1), 8),
            ('ROWBACKGROUNDS',(0,1),(-1,-1), [colors.white, lilasB]),
            ('GRID',          (0,0),(-1,-1), 0.4, borda),
            ('TOPPADDING',    (0,0),(-1,-1), 4),
            ('BOTTOMPADDING', (0,0),(-1,-1), 4),
            ('LEFTPADDING',   (0,0),(-1,-1), 5),
            ('RIGHTPADDING',  (0,0),(-1,-1), 5),
            ('VALIGN',        (0,0),(-1,-1), 'MIDDLE'),
        ])
        for i, ch in enumerate(chamados, 1):
            if ch.status == 'aberto':
                ts_ch.add('TEXTCOLOR', (3,i),(3,i), colors.HexColor('#dc2626'))
                ts_ch.add('FONTNAME',  (3,i),(3,i), 'Helvetica-Bold')
            else:
                ts_ch.add('TEXTCOLOR', (3,i),(3,i), colors.HexColor('#16a34a'))
        t_ch.setStyle(ts_ch)
        elems.append(t_ch)
    else:
        elems.append(Paragraph("Nenhum chamado registrado.", cs))
    elems.append(Spacer(1, 16))

    # Histórico de movimentações
    if movs:
        elems.append(Paragraph("Histórico de Movimentações",
            ParagraphStyle('h2', fontSize=11, fontName='Helvetica-Bold', textColor=roxoL, spaceBefore=8, spaceAfter=6)))
        rows_mv = [['Data','Campo','Valor Anterior','Valor Novo','Usuário']]
        for mv in movs:
            rows_mv.append([P(mv.criado_em.strftime('%d/%m/%Y %H:%M') if mv.criado_em else ''),
                             P(mv.campo_alterado, True), P(mv.valor_anterior), P(mv.valor_novo), P(mv.usuario.nome if mv.usuario else '—')])
        t_mv = Table(rows_mv, colWidths=[100,80,100,100,80], repeatRows=1)
        t_mv.setStyle(TableStyle([
            ('BACKGROUND',    (0,0),(-1,0), roxo),
            ('TEXTCOLOR',     (0,0),(-1,0), colors.white),
            ('FONTNAME',      (0,0),(-1,0), 'Helvetica-Bold'),
            ('FONTSIZE',      (0,0),(-1,-1), 8),
            ('ROWBACKGROUNDS',(0,1),(-1,-1), [colors.white, lilasB]),
            ('GRID',          (0,0),(-1,-1), 0.4, borda),
            ('TOPPADDING',    (0,0),(-1,-1), 4),
            ('BOTTOMPADDING', (0,0),(-1,-1), 4),
            ('LEFTPADDING',   (0,0),(-1,-1), 5),
            ('RIGHTPADDING',  (0,0),(-1,-1), 5),
            ('VALIGN',        (0,0),(-1,-1), 'MIDDLE'),
        ]))
        elems.append(t_mv)

    doc.build(elems)
    out.seek(0)
    pat = eq.numero_patrimonio.replace('/', '-')
    return send_file(out, mimetype='application/pdf', as_attachment=False,
                     download_name=f'ficha_{pat}_{date.today()}.pdf')

# ══════════════════════════════════════════════════════════════════════════════
# ALERTAS DE VENCIMENTO
# ══════════════════════════════════════════════════════════════════════════════

def enviar_alertas_vencimento():
    with app.app_context():
        cfg = ConfigAlerta.query.first()
        if not cfg or not cfg.ativo or not cfg.get_emails():
            return
        hoje  = date.today()
        aviso = cfg.dias_aviso or 10
        alvo  = hoje + timedelta(days=aviso)
        alertas = []
        vencidos = []
        for Model, tipo in [(Monitor,"Monitor"),(Desktop,"Desktop"),(Estabilizador,"Estabilizador"),(Notebook,"Notebook")]:
            for eq in Model.query.filter_by(ativo=True).filter(Model.data_fim_fidelidade == alvo).all():
                alertas.append({"tipo":tipo,"pat":eq.numero_patrimonio,
                    "pa":eq.pa or "—",
                    "empresa":eq.empresa.nome if eq.empresa else "",
                    "local":eq.localizacao,
                    "data":eq.data_fim_fidelidade.strftime("%d/%m/%Y")})
            for eq in Model.query.filter_by(ativo=True).filter(Model.data_fim_fidelidade < hoje).all():
                vencidos.append({"tipo":tipo,"pat":eq.numero_patrimonio,
                    "pa":eq.pa or "—",
                    "empresa":eq.empresa.nome if eq.empresa else "",
                    "local":eq.localizacao,
                    "data":eq.data_fim_fidelidade.strftime("%d/%m/%Y")})
        if not alertas and not vencidos:
            return
        def tabela(rows, cor_txt, cor_bg):
            header = ("<table style='width:100%;border-collapse:collapse;font-size:13px;margin-bottom:16px;'>"
                      "<thead><tr style='background:#3d1236;color:white;'>"
                      "<th style='padding:8px;'>Tipo</th>"
                      "<th style='padding:8px;'>Patrimônio</th>"
                      "<th style='padding:8px;'>PA</th>"
                      "<th style='padding:8px;'>Empresa</th>"
                      "<th style='padding:8px;'>Localização</th>"
                      "<th style='padding:8px;'>Fim Fidelidade</th>"
                      "</tr></thead><tbody>")
            linhas = "".join(
                "<tr style='background:{bg};'>"
                "<td style='padding:8px;border:1px solid #e5e7eb;'>{tipo}</td>"
                "<td style='padding:8px;border:1px solid #e5e7eb;font-weight:700;'>{pat}</td>"
                "<td style='padding:8px;border:1px solid #e5e7eb;'>{pa}</td>"
                "<td style='padding:8px;border:1px solid #e5e7eb;'>{empresa}</td>"
                "<td style='padding:8px;border:1px solid #e5e7eb;'>{local}</td>"
                "<td style='padding:8px;border:1px solid #e5e7eb;color:{cor};font-weight:700;'>{data}</td>"
                "</tr>".format(bg=cor_bg, cor=cor_txt, **r) for r in rows)
            return header + linhas + "</tbody></table>"
        corpo = (
            "<div style='font-family:Arial,sans-serif;max-width:700px;margin:0 auto;'>"
            "<div style='background:#3d1236;padding:24px;border-radius:12px 12px 0 0;'>"
            "<h1 style='color:white;margin:0;font-size:20px;'>⚠ Alerta de Vencimento de Contratos</h1>"
            "<p style='color:rgba(255,255,255,.7);margin:4px 0 0;'>PatrimônioTech — Fama Soluções</p>"
            "</div>"
            "<div style='background:white;padding:24px;border:1px solid #e5e7eb;border-top:none;'>"
        )
        if alertas:
            corpo += ("<h3 style='color:#d97706;margin:0 0 12px;'>⏰ Vencem em " + str(aviso) + " dias</h3>"
                      + tabela(alertas, "#d97706", "#fffbeb"))
        if vencidos:
            corpo += ("<h3 style='color:#dc2626;margin:16px 0 12px;'>🔴 Contratos vencidos</h3>"
                      + tabela(vencidos, "#dc2626", "#fef2f2"))
        corpo += ("<p style='color:#6b7280;font-size:12px;margin-top:20px;'>"
                  "Gerado automaticamente em " + hoje.strftime("%d/%m/%Y") + " pelo PatrimônioTech.</p>"
                  "</div></div>")
        total = len(alertas) + len(vencidos)
        assunto = "⚠ PatrimônioTech — {} contrato(s) precisam de atenção".format(total)
        for email in cfg.get_emails():
            enviar_email(email, assunto, corpo)
        cfg.ultimo_envio = datetime.utcnow()
        db.session.commit()
        print("[ALERTA] Enviado: {} a vencer, {} vencidos.".format(len(alertas), len(vencidos)))

@app.route('/api/config-alerta', methods=['GET'])
@perfil_required('admin')
def get_config_alerta():
    cfg = ConfigAlerta.query.first()
    if not cfg:
        cfg = ConfigAlerta(); db.session.add(cfg); db.session.commit()
    return jsonify({'emails': cfg.emails or '', 'dias_aviso': cfg.dias_aviso or 10,
                    'ativo': cfg.ativo,
                    'ultimo_envio': cfg.ultimo_envio.strftime('%d/%m/%Y %H:%M') if cfg.ultimo_envio else ''})

@app.route('/api/config-alerta', methods=['POST'])
@perfil_required('admin')
def salvar_config_alerta():
    data = request.json
    cfg  = ConfigAlerta.query.first()
    if not cfg:
        cfg = ConfigAlerta(); db.session.add(cfg)
    cfg.emails     = data.get('emails', '')
    cfg.dias_aviso = int(data.get('dias_aviso', 10))
    cfg.ativo      = data.get('ativo', True)
    db.session.commit()
    return jsonify({'sucesso': True})

@app.route('/api/config-alerta/testar', methods=['POST'])
@perfil_required('admin')
def testar_alerta():
    cfg = ConfigAlerta.query.first()
    if not cfg or not cfg.get_emails():
        return jsonify({'sucesso': False, 'erro': 'Nenhum e-mail cadastrado.'})
    ok, erro = enviar_email(cfg.get_emails()[0],
        '✅ PatrimônioTech — Teste de alerta de vencimento',
        '<div style="font-family:Arial;padding:24px;max-width:500px;">'
        '<div style="background:#3d1236;padding:20px;border-radius:10px 10px 0 0;">'
        '<h2 style="color:white;margin:0;">✅ Alerta configurado!</h2></div>'
        '<div style="background:white;padding:20px;border:1px solid #e5e7eb;">'
        '<p>Os alertas de vencimento de contratos estão funcionando corretamente.</p>'
        '<p style="color:#6b7280;font-size:12px;">PatrimônioTech — Fama Soluções</p>'
        '</div></div>')
    if ok:
        return jsonify({'sucesso': True, 'mensagem': 'E-mail de teste enviado para {}'.format(cfg.get_emails()[0])})
    return jsonify({'sucesso': False, 'erro': erro})

@app.route('/api/config-alerta/disparar', methods=['POST'])
@perfil_required('admin')
def disparar_alerta():
    import threading
    threading.Thread(target=enviar_alertas_vencimento, daemon=True).start()
    return jsonify({'sucesso': True, 'mensagem': 'Verificação iniciada em segundo plano.'})


def alerta_chamados_abertos():
    """Notifica sobre chamados abertos há mais de X dias (padrão 7)."""
    with app.app_context():
        cfg = ConfigAlerta.query.first()
        if not cfg or not cfg.ativo or not cfg.get_emails():
            return
        limite_dias = 7
        corte = datetime.utcnow() - timedelta(days=limite_dias)
        velhos = (Chamado.query.filter_by(status='aberto')
                  .filter(Chamado.criado_em <= corte).all())
        if not velhos:
            return
        linhas = ""
        for ch in velhos:
            eq = _get_equipamento(ch.tipo_equipamento, ch.equipamento_id)
            dias_aberto = (date.today() - ch.data_abertura).days if ch.data_abertura else '?'
            pat = eq.numero_patrimonio if eq else '—'
            emp = eq.empresa.nome if eq and eq.empresa else '—'
            linhas += (
                "<tr style='background:#fffbeb;'>"
                f"<td style='padding:8px;border:1px solid #e5e7eb;'>{ch.numero_chamado}</td>"
                f"<td style='padding:8px;border:1px solid #e5e7eb;'>{ch.tipo_equipamento}</td>"
                f"<td style='padding:8px;border:1px solid #e5e7eb;font-weight:700;'>{pat}</td>"
                f"<td style='padding:8px;border:1px solid #e5e7eb;'>{emp}</td>"
                f"<td style='padding:8px;border:1px solid #e5e7eb;color:#d97706;font-weight:700;'>{dias_aberto} dias</td>"
                "</tr>")
        corpo = (
            "<div style='font-family:Arial,sans-serif;max-width:700px;margin:0 auto;'>"
            "<div style='background:#3d1236;padding:24px;border-radius:12px 12px 0 0;'>"
            "<h1 style='color:white;margin:0;font-size:20px;'>🔧 Chamados em aberto há muito tempo</h1>"
            "<p style='color:rgba(255,255,255,.7);margin:4px 0 0;'>PatrimônioTech — Fama Soluções</p>"
            "</div>"
            "<div style='background:white;padding:24px;border:1px solid #e5e7eb;border-top:none;'>"
            f"<p style='margin:0 0 14px;font-size:13px;'>{len(velhos)} chamado(s) estão abertos há mais de {limite_dias} dias:</p>"
            "<table style='width:100%;border-collapse:collapse;font-size:13px;'>"
            "<thead><tr style='background:#3d1236;color:white;'>"
            "<th style='padding:8px;'>Nº Chamado</th><th style='padding:8px;'>Tipo</th>"
            "<th style='padding:8px;'>Patrimônio</th><th style='padding:8px;'>Empresa</th>"
            "<th style='padding:8px;'>Dias em aberto</th>"
            "</tr></thead><tbody>" + linhas + "</tbody></table>"
            f"<p style='color:#6b7280;font-size:12px;margin-top:20px;'>Gerado em {date.today().strftime('%d/%m/%Y')}.</p>"
            "</div></div>")
        for email in cfg.get_emails():
            enviar_email(email, f"🔧 PatrimônioTech — {len(velhos)} chamado(s) abertos há mais de {limite_dias} dias", corpo)
        print(f"[CHAMADOS] Alerta enviado: {len(velhos)} chamados antigos.")

# ══════════════════════════════════════════════════════════════════════════════
# STARTUP
# ══════════════════════════════════════════════════════════════════════════════

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
        # Cria admin padrão se não existir
        if not Usuario.query.filter_by(email='admin@patrimonio.local').first():
            admin = Usuario(nome='Administrador', email='admin@patrimonio.local',
                            perfil='admin', senha_hash=hash_senha('Admin@123'))
            db.session.add(admin); db.session.commit()
            print("Admin padrão criado: admin@patrimonio.local / Admin@123")
        print("Tabelas criadas/verificadas.")

    # Scheduler de alertas diários às 08:00
    if HAS_SCHEDULER:
        _scheduler = BackgroundScheduler(timezone='America/Fortaleza')
        _scheduler.add_job(enviar_alertas_vencimento, 'cron', hour=8, minute=0,
                           id='alerta_vencimento', replace_existing=True)
        _scheduler.start()
        print('Scheduler ativo — alertas enviados diariamente às 08:00.')
    else:
        print('APScheduler não instalado. Execute: pip install apscheduler --break-system-packages')

    ssl_cert = os.getenv('SSL_CERT')
    ssl_key  = os.getenv('SSL_KEY')
    if ssl_cert and ssl_key and os.path.exists(ssl_cert) and os.path.exists(ssl_key):
        print("SSL ativo — acesse: https://0.0.0.0:5000")
        app.run(host='0.0.0.0', port=5000, debug=True, ssl_context=(ssl_cert, ssl_key))
    else:
        print("Servidor iniciado em http://0.0.0.0:5000")
        app.run(host='0.0.0.0', port=5000, debug=True)
